#!/usr/bin/env python3
"""
KaBuM Brazil — Market Intelligence Dashboard
Cooler Master Competitive Analysis — February 2026
"""

import csv
import os
import shutil
import statistics
from collections import defaultdict
from copy import copy
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, NamedStyle,
    PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule

# ── Paths ──
CSV_PATH = "/Users/viniciuslimadalbem/.openclaw/workspace-pesquisa/kabum-scraper/data/historico/kabum_2026-02-16_full_clean.csv"
OUT_PATHS = [
    "/Users/viniciuslimadalbem/.openclaw/workspace-pesquisa/kabum-scraper/data/historico/KaBuM_Dashboard_2026-02-16.xlsx",
    os.path.expanduser("~/Desktop/KaBuM_Dashboard_2026-02-16.xlsx"),
    os.path.expanduser("~/Library/CloudStorage/OneDrive-CoolerMaster/2026/Jubinha/kabum-scraper/historico/KaBuM_Dashboard_2026-02-16.xlsx"),
]

# ── Colors ──
DARK_HEADER = "1F2937"
SUB_HEADER = "374151"
ACCENT_BLUE = "3B82F6"
SUCCESS = "10B981"
WARNING = "F59E0B"
DANGER = "EF4444"
ALT_ROW = "F9FAFB"
CM_HIGHLIGHT = "FEF3C7"
WHITE = "FFFFFF"
LIGHT_BLUE = "DBEAFE"
LIGHT_GREEN = "D1FAE5"

# ── Fills ──
fill_header = PatternFill("solid", fgColor=DARK_HEADER)
fill_subheader = PatternFill("solid", fgColor=SUB_HEADER)
fill_accent = PatternFill("solid", fgColor=ACCENT_BLUE)
fill_alt = PatternFill("solid", fgColor=ALT_ROW)
fill_cm = PatternFill("solid", fgColor=CM_HIGHLIGHT)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_success = PatternFill("solid", fgColor=LIGHT_GREEN)
fill_light_blue = PatternFill("solid", fgColor=LIGHT_BLUE)
fill_danger_light = PatternFill("solid", fgColor="FEE2E2")
fill_warning_light = PatternFill("solid", fgColor="FEF3C7")

# ── Fonts ──
font_header = Font(bold=True, color=WHITE, size=11, name="Calibri")
font_subheader = Font(bold=True, color=WHITE, size=10, name="Calibri")
font_title = Font(bold=True, color=DARK_HEADER, size=20, name="Calibri")
font_subtitle = Font(bold=True, color=SUB_HEADER, size=14, name="Calibri")
font_kpi_value = Font(bold=True, color=ACCENT_BLUE, size=24, name="Calibri")
font_kpi_label = Font(bold=True, color=SUB_HEADER, size=10, name="Calibri")
font_normal = Font(size=10, name="Calibri")
font_bold = Font(bold=True, size=10, name="Calibri")
font_cm = Font(bold=True, size=10, name="Calibri", color="92400E")
font_small = Font(size=9, name="Calibri", color="6B7280")

# ── Borders ──
thin_border = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
bottom_border = Border(bottom=Side(style="medium", color=ACCENT_BLUE))

# ── Alignments ──
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

# ── Category names ──
CAT_MAP = {"fontes": "PSU", "gabinetes": "Cases", "coolers": "Coolers"}
CAT_MAP_REV = {v: k for k, v in CAT_MAP.items()}

# ── Price segments ──
def price_segment(price):
    if price < 200: return "Budget (<R$200)"
    elif price < 500: return "Mid (R$200-500)"
    elif price < 1000: return "Premium (R$500-1K)"
    else: return "Ultra-Premium (R$1K+)"

SEGMENTS_ORDER = ["Budget (<R$200)", "Mid (R$200-500)", "Premium (R$500-1K)", "Ultra-Premium (R$1K+)"]


def safe_float(v, default=0.0):
    try:
        return float(v) if v else default
    except (ValueError, TypeError):
        return default

def safe_int(v, default=0):
    try:
        return int(float(v)) if v else default
    except (ValueError, TypeError):
        return default


# ── Load Data ──
def load_data():
    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            # Normalize
            cat = row.get("categoria", row.get("\ufeffcategoria", ""))
            r = {
                "categoria": cat,
                "cat_en": CAT_MAP.get(cat, cat),
                "id": row.get("id", ""),
                "nome": row.get("nome", ""),
                "marca": row.get("marca", ""),
                "partnumber": row.get("partnumber", ""),
                "url": row.get("url", ""),
                "preco_brl": safe_float(row.get("preco_brl")),
                "preco_pix": safe_float(row.get("preco_pix")),
                "preco_lista": safe_float(row.get("preco_lista")),
                "preco_com_desconto": safe_float(row.get("preco_com_desconto")),
                "desconto_pct": safe_float(row.get("desconto_pct")),
                "vendas_estimadas": safe_int(row.get("vendas_estimadas")),
                "estoque": safe_int(row.get("estoque")),
                "disponivel": row.get("disponivel", "") == "True",
                "is_cm": row.get("is_cooler_master", "") == "True",
                "avaliacao": safe_float(row.get("avaliacao")),
                "num_avaliacoes": safe_int(row.get("num_avaliacoes")),
                "warranty": row.get("warranty", ""),
                "subcategoria": row.get("subcategoria", ""),
                "frete_gratis": row.get("frete_gratis", "") == "Sim",
                "is_prime": row.get("is_prime", "") == "Sim",
                "seller_type": row.get("seller_type", ""),
                "data_coleta": row.get("data_coleta", ""),
                "spec_potencia": row.get("spec_potência", ""),
                "spec_tipo": row.get("spec_tipo", ""),
                "spec_tdp": row.get("spec_tdp", ""),
                "spec_tamanho": row.get("spec_tamanho", ""),
                "spec_cor": row.get("spec_cor", ""),
                "spec_eficiencia": row.get("spec_eficiência", ""),
                "spec_certificacao": row.get("spec_certificação", ""),
            }
            r["revenue_est"] = r["preco_pix"] * r["vendas_estimadas"]
            rows.append(r)
    return rows


# ── Helpers ──
def style_header_row(ws, row_num, max_col, fill=None, font=None):
    f = fill or fill_header
    fn = font or font_header
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = f
        cell.font = fn
        cell.alignment = align_center
        cell.border = thin_border

def write_row(ws, row_num, values, is_cm=False, bold=False):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = font_cm if is_cm else (font_bold if bold else font_normal)
        cell.border = thin_border
        cell.alignment = align_center if col > 1 else align_left
        if is_cm:
            cell.fill = fill_cm
        elif row_num % 2 == 0:
            cell.fill = fill_alt

def write_section_title(ws, row, col, title, width=4):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width - 1)
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = font_subtitle
    cell.alignment = align_left
    cell.border = bottom_border

def write_kpi_box(ws, row, col, value, label, width=2, height=3):
    # Merge for value
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + width - 1)
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font_kpi_value
    cell.alignment = align_center
    # Border the box
    for r in range(row, row + height):
        for c in range(col, col + width):
            ws.cell(row=r, column=c).border = Border(
                left=Side(style="medium", color=ACCENT_BLUE) if c == col else Side(style="thin", color="E5E7EB"),
                right=Side(style="medium", color=ACCENT_BLUE) if c == col + width - 1 else Side(style="thin", color="E5E7EB"),
                top=Side(style="medium", color=ACCENT_BLUE) if r == row else Side(style="thin", color="E5E7EB"),
                bottom=Side(style="medium", color=ACCENT_BLUE) if r == row + height - 1 else Side(style="thin", color="E5E7EB"),
            )
            ws.cell(row=r, column=c).fill = fill_white
    # Label
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + width - 1)
    lbl = ws.cell(row=row + 2, column=col, value=label)
    lbl.font = font_kpi_label
    lbl.alignment = align_center

def auto_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        max_len = min_w
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[col_letter].width = max_len + 2

def add_filters(ws, row, max_col):
    ws.auto_filter.ref = f"A{row}:{get_column_letter(max_col)}{ws.max_row}"

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def fmt_brl(cell):
    cell.number_format = '#,##0.00'

def fmt_pct(cell):
    cell.number_format = '0.0%'

def fmt_int(cell):
    cell.number_format = '#,##0'


# ═══════════════════════════════════════════════
#  TAB BUILDERS
# ═══════════════════════════════════════════════

def build_executive_summary(wb, data):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = ACCENT_BLUE

    # Title
    ws.merge_cells("A1:L1")
    t = ws.cell(row=1, column=1, value="KaBuM Brazil — Market Intelligence Report")
    t.font = font_title
    t.alignment = align_center
    t.fill = PatternFill("solid", fgColor=DARK_HEADER)
    t.font = Font(bold=True, color=WHITE, size=22, name="Calibri")

    ws.merge_cells("A2:L2")
    s = ws.cell(row=2, column=1, value="Cooler Master Competitive Analysis — February 2026")
    s.font = Font(bold=True, color=ACCENT_BLUE, size=14, name="Calibri")
    s.alignment = align_center

    ws.merge_cells("A3:L3")
    d = ws.cell(row=3, column=1, value=f"Data Collection: {data[0]['data_coleta']} | Source: KaBuM.com.br | Total Products: {len(data)}")
    d.font = font_small
    d.alignment = align_center

    # KPIs
    cm_products = [r for r in data if r["is_cm"]]
    total = len(data)
    cats = set(r["categoria"] for r in data)
    cm_count = len(cm_products)
    
    prices_all = [r["preco_pix"] for r in data if r["preco_pix"] > 0]
    prices_cm = [r["preco_pix"] for r in cm_products if r["preco_pix"] > 0]
    avg_market = statistics.mean(prices_all) if prices_all else 0
    avg_cm = statistics.mean(prices_cm) if prices_cm else 0

    # KPI boxes row 5
    row = 5
    write_kpi_box(ws, row, 1, total, "Products Monitored")
    write_kpi_box(ws, row, 3, len(cats), "Categories")
    write_kpi_box(ws, row, 5, cm_count, "CM Products")
    write_kpi_box(ws, row, 7, f"R$ {avg_cm:,.0f}", "CM Avg Price")
    write_kpi_box(ws, row, 9, f"R$ {avg_market:,.0f}", "Market Avg Price")
    
    # CM share by category
    row = 9
    cm_rev = sum(r["revenue_est"] for r in cm_products)
    total_rev = sum(r["revenue_est"] for r in data)
    rev_share = cm_rev / total_rev * 100 if total_rev else 0
    write_kpi_box(ws, row, 1, f"{cm_count/total*100:.1f}%", "CM SKU Share")
    write_kpi_box(ws, row, 3, f"{rev_share:.1f}%", "CM Revenue Share")
    
    # Price positioning
    cm_segs = defaultdict(int)
    for p in cm_products:
        cm_segs[price_segment(p["preco_pix"])] += 1
    seg_str = ", ".join(f"{s}: {cm_segs.get(s,0)}" for s in SEGMENTS_ORDER)
    write_kpi_box(ws, row, 5, seg_str, "CM Price Positioning", width=6)

    # CM Products table
    row = 13
    write_section_title(ws, row, 1, "Cooler Master Products on KaBuM", width=10)
    row += 1
    headers = ["Product Name", "Category", "Price (PIX)", "List Price", "Discount %",
               "Rating", "Reviews", "Est. Sales", "Stock", "Available"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    
    for p in sorted(cm_products, key=lambda x: x["preco_pix"], reverse=True):
        row += 1
        vals = [p["nome"], p["cat_en"], p["preco_pix"], p["preco_lista"],
                p["desconto_pct"] / 100 if p["desconto_pct"] else 0,
                p["avaliacao"], p["num_avaliacoes"], p["vendas_estimadas"],
                p["estoque"], "Yes" if p["disponivel"] else "No"]
        write_row(ws, row, vals, is_cm=True)
        fmt_brl(ws.cell(row=row, column=3))
        fmt_brl(ws.cell(row=row, column=4))
        fmt_pct(ws.cell(row=row, column=5))

    # CM Share by Category mini-table
    row += 2
    write_section_title(ws, row, 1, "CM Market Share by Category", width=6)
    row += 1
    headers2 = ["Category", "Total SKUs", "CM SKUs", "CM SKU %", "CM Avg Price", "Category Avg Price"]
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers2))
    
    for cat_pt, cat_en in CAT_MAP.items():
        row += 1
        cat_all = [r for r in data if r["categoria"] == cat_pt]
        cat_cm = [r for r in cat_all if r["is_cm"]]
        cat_prices = [r["preco_pix"] for r in cat_all if r["preco_pix"] > 0]
        cm_prices_cat = [r["preco_pix"] for r in cat_cm if r["preco_pix"] > 0]
        vals = [
            cat_en, len(cat_all), len(cat_cm),
            len(cat_cm) / len(cat_all) if cat_all else 0,
            statistics.mean(cm_prices_cat) if cm_prices_cat else 0,
            statistics.mean(cat_prices) if cat_prices else 0,
        ]
        write_row(ws, row, vals, is_cm=True)
        fmt_pct(ws.cell(row=row, column=4))
        fmt_brl(ws.cell(row=row, column=5))
        fmt_brl(ws.cell(row=row, column=6))

    auto_width(ws)
    ws.sheet_view.showGridLines = False
    freeze(ws, "A5")
    print(f"  Executive Summary: rows up to {row}")


def build_market_overview(wb, data):
    ws = wb.create_sheet("Market Overview")
    ws.sheet_properties.tabColor = ACCENT_BLUE

    row = 1
    write_section_title(ws, row, 1, "Market Overview — KaBuM Hardware", width=8)

    # Products by category
    row = 3
    headers = ["Category", "Products", "% of Total", "Brands", "Min Price", "Max Price",
               "Avg Price", "Median Price", "Total Est. Revenue"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    for cat_pt, cat_en in CAT_MAP.items():
        row += 1
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        prices = [r["preco_pix"] for r in cat_d if r["preco_pix"] > 0]
        brands = set(r["marca"] for r in cat_d)
        rev = sum(r["revenue_est"] for r in cat_d)
        vals = [
            cat_en, len(cat_d), len(cat_d)/len(data),
            len(brands),
            min(prices) if prices else 0, max(prices) if prices else 0,
            statistics.mean(prices) if prices else 0,
            statistics.median(prices) if prices else 0,
            rev,
        ]
        write_row(ws, row, vals)
        fmt_pct(ws.cell(row=row, column=3))
        for c in [5,6,7,8,9]: fmt_brl(ws.cell(row=row, column=c))

    # Top 10 brands by SKUs
    row += 2
    write_section_title(ws, row, 1, "Top 10 Brands by SKU Count", width=6)
    row += 1
    brand_counts = defaultdict(int)
    brand_rev = defaultdict(float)
    brand_rating = defaultdict(list)
    for r_d in data:
        brand_counts[r_d["marca"]] += 1
        brand_rev[r_d["marca"]] += r_d["revenue_est"]
        if r_d["avaliacao"] > 0:
            brand_rating[r_d["marca"]].append(r_d["avaliacao"])

    headers2 = ["Rank", "Brand", "SKUs", "% Share", "Avg Rating", "Est. Revenue"]
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers2))

    for i, (brand, cnt) in enumerate(sorted(brand_counts.items(), key=lambda x: -x[1])[:10], 1):
        row += 1
        avg_rat = statistics.mean(brand_rating[brand]) if brand_rating[brand] else 0
        is_cm = brand.lower() == "cooler master"
        write_row(ws, row, [i, brand, cnt, cnt/len(data), avg_rat, brand_rev[brand]], is_cm=is_cm)
        fmt_pct(ws.cell(row=row, column=4))
        fmt_brl(ws.cell(row=row, column=6))

    # Top 10 by revenue
    row += 2
    write_section_title(ws, row, 1, "Top 10 Brands by Estimated Revenue", width=6)
    row += 1
    for c, h in enumerate(["Rank", "Brand", "Est. Revenue", "Revenue %", "SKUs", "Avg Price"], 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 6)

    total_rev = sum(brand_rev.values())
    brand_prices = defaultdict(list)
    for r_d in data:
        if r_d["preco_pix"] > 0:
            brand_prices[r_d["marca"]].append(r_d["preco_pix"])

    for i, (brand, rev) in enumerate(sorted(brand_rev.items(), key=lambda x: -x[1])[:10], 1):
        row += 1
        is_cm = brand.lower() == "cooler master"
        avg_p = statistics.mean(brand_prices[brand]) if brand_prices[brand] else 0
        write_row(ws, row, [i, brand, rev, rev/total_rev if total_rev else 0,
                           brand_counts[brand], avg_p], is_cm=is_cm)
        fmt_brl(ws.cell(row=row, column=3))
        fmt_pct(ws.cell(row=row, column=4))
        fmt_brl(ws.cell(row=row, column=6))

    # Stock distribution
    row += 2
    write_section_title(ws, row, 1, "Stock Distribution by Category", width=5)
    row += 1
    for c, h in enumerate(["Category", "Total Stock", "Avg Stock/SKU", "Min", "Max"], 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 5)

    for cat_pt, cat_en in CAT_MAP.items():
        row += 1
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        stocks = [r["estoque"] for r in cat_d]
        write_row(ws, row, [cat_en, sum(stocks), statistics.mean(stocks) if stocks else 0,
                           min(stocks) if stocks else 0, max(stocks) if stocks else 0])
        for c in [2,3,4,5]: fmt_int(ws.cell(row=row, column=c))

    auto_width(ws)
    freeze(ws, "A4")
    print(f"  Market Overview: rows up to {row}")


def build_cm_analysis(wb, data):
    ws = wb.create_sheet("Cooler Master Analysis")
    ws.sheet_properties.tabColor = "92400E"

    cm = [r for r in data if r["is_cm"]]
    
    row = 1
    write_section_title(ws, row, 1, "Cooler Master — Full Product Analysis", width=14)

    row = 3
    headers = ["Product Name", "Category", "SKU", "Price (PIX)", "List Price", "Discount %",
               "vs Cat. Avg", "vs Cat. Median", "Price Rank in Cat.",
               "Rating", "Cat. Avg Rating", "Reviews", "Est. Sales",
               "Stock", "Cat. Avg Stock", "Warranty"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    # Precompute category stats
    cat_stats = {}
    for cat_pt in CAT_MAP:
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        prices = sorted([r["preco_pix"] for r in cat_d if r["preco_pix"] > 0])
        ratings = [r["avaliacao"] for r in cat_d if r["avaliacao"] > 0]
        stocks = [r["estoque"] for r in cat_d]
        cat_stats[cat_pt] = {
            "avg_price": statistics.mean(prices) if prices else 0,
            "med_price": statistics.median(prices) if prices else 0,
            "prices_sorted": prices,
            "avg_rating": statistics.mean(ratings) if ratings else 0,
            "avg_stock": statistics.mean(stocks) if stocks else 0,
        }

    for p in sorted(cm, key=lambda x: x["categoria"]):
        row += 1
        cs = cat_stats[p["categoria"]]
        # Price rank
        rank = 1
        for pp in cs["prices_sorted"]:
            if pp < p["preco_pix"]:
                rank += 1
        rank_total = len(cs["prices_sorted"])
        
        vs_avg = (p["preco_pix"] - cs["avg_price"]) / cs["avg_price"] * 100 if cs["avg_price"] else 0
        vs_med = (p["preco_pix"] - cs["med_price"]) / cs["med_price"] * 100 if cs["med_price"] else 0
        
        vals = [
            p["nome"], p["cat_en"], p["partnumber"],
            p["preco_pix"], p["preco_lista"],
            p["desconto_pct"] / 100 if p["desconto_pct"] else 0,
            vs_avg / 100, vs_med / 100,
            f"{rank}/{rank_total}",
            p["avaliacao"], cs["avg_rating"], p["num_avaliacoes"],
            p["vendas_estimadas"], p["estoque"], cs["avg_stock"],
            p["warranty"],
        ]
        write_row(ws, row, vals, is_cm=True)
        fmt_brl(ws.cell(row=row, column=4))
        fmt_brl(ws.cell(row=row, column=5))
        fmt_pct(ws.cell(row=row, column=6))
        fmt_pct(ws.cell(row=row, column=7))
        fmt_pct(ws.cell(row=row, column=8))
        fmt_int(ws.cell(row=row, column=14))

    # Gap Analysis
    row += 2
    write_section_title(ws, row, 1, "Gap Analysis — Where CM Has No Products", width=8)
    row += 1
    headers2 = ["Category", "Segment", "Market Products", "Top Brand", "Top Brand SKUs", "Opportunity"]
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers2))

    for cat_pt, cat_en in CAT_MAP.items():
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        cm_cat = [r for r in cat_d if r["is_cm"]]
        cm_segs = set(price_segment(r["preco_pix"]) for r in cm_cat)
        for seg in SEGMENTS_ORDER:
            seg_products = [r for r in cat_d if price_segment(r["preco_pix"]) == seg]
            if seg_products and seg not in cm_segs:
                row += 1
                brand_count = defaultdict(int)
                for r_d in seg_products:
                    brand_count[r_d["marca"]] += 1
                top_brand = max(brand_count, key=brand_count.get)
                write_row(ws, row, [cat_en, seg, len(seg_products), top_brand,
                                   brand_count[top_brand], "NO CM PRODUCT"])
                ws.cell(row=row, column=6).font = Font(bold=True, color=DANGER, size=10)

    auto_width(ws)
    freeze(ws, "A4")
    add_filters(ws, 3, len(headers))
    print(f"  CM Analysis: rows up to {row}")


def build_price_positioning(wb, data):
    ws = wb.create_sheet("Price Positioning")
    ws.sheet_properties.tabColor = WARNING

    row = 1
    write_section_title(ws, row, 1, "Price Segmentation Analysis", width=10)

    # Overall segmentation
    row = 3
    headers = ["Segment", "Total Products", "% of Total", "Avg Price", "Avg Rating",
               "CM Products", "Top 3 Brands"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    for seg in SEGMENTS_ORDER:
        row += 1
        seg_d = [r for r in data if price_segment(r["preco_pix"]) == seg]
        cm_seg = [r for r in seg_d if r["is_cm"]]
        prices = [r["preco_pix"] for r in seg_d if r["preco_pix"] > 0]
        ratings = [r["avaliacao"] for r in seg_d if r["avaliacao"] > 0]
        brand_c = defaultdict(int)
        for r_d in seg_d:
            brand_c[r_d["marca"]] += 1
        top3 = ", ".join(b for b, _ in sorted(brand_c.items(), key=lambda x: -x[1])[:3])
        write_row(ws, row, [seg, len(seg_d), len(seg_d)/len(data),
                           statistics.mean(prices) if prices else 0,
                           statistics.mean(ratings) if ratings else 0,
                           len(cm_seg), top3], is_cm=bool(cm_seg))
        fmt_pct(ws.cell(row=row, column=3))
        fmt_brl(ws.cell(row=row, column=4))

    # Per-category segmentation
    for cat_pt, cat_en in CAT_MAP.items():
        row += 2
        write_section_title(ws, row, 1, f"Price Positioning — {cat_en}", width=8)
        row += 1
        headers2 = ["Brand", "Budget", "Mid", "Premium", "Ultra-Premium", "Total", "Avg Price"]
        for c, h in enumerate(headers2, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers2))

        cat_d = [r for r in data if r["categoria"] == cat_pt]
        brand_seg = defaultdict(lambda: defaultdict(int))
        brand_total = defaultdict(int)
        brand_avg = defaultdict(list)
        for r_d in cat_d:
            brand_seg[r_d["marca"]][price_segment(r_d["preco_pix"])] += 1
            brand_total[r_d["marca"]] += 1
            if r_d["preco_pix"] > 0:
                brand_avg[r_d["marca"]].append(r_d["preco_pix"])

        for brand in sorted(brand_total, key=lambda b: -brand_total[b])[:15]:
            row += 1
            is_cm = brand.lower() == "cooler master"
            write_row(ws, row, [
                brand,
                brand_seg[brand].get("Budget (<R$200)", 0),
                brand_seg[brand].get("Mid (R$200-500)", 0),
                brand_seg[brand].get("Premium (R$500-1K)", 0),
                brand_seg[brand].get("Ultra-Premium (R$1K+)", 0),
                brand_total[brand],
                statistics.mean(brand_avg[brand]) if brand_avg[brand] else 0,
            ], is_cm=is_cm)
            fmt_brl(ws.cell(row=row, column=7))

    auto_width(ws)
    freeze(ws, "A4")
    print(f"  Price Positioning: rows up to {row}")


def build_competitive_tab(wb, data, cat_pt, cat_en, key_competitors):
    ws = wb.create_sheet(f"Competitive — {cat_en}")
    ws.sheet_properties.tabColor = ACCENT_BLUE

    cat_d = [r for r in data if r["categoria"] == cat_pt]
    
    row = 1
    write_section_title(ws, row, 1, f"Competitive Analysis — {cat_en}", width=12)

    # Brand ranking
    row = 3
    headers = ["Rank", "Brand", "SKUs", "Avg Price (PIX)", "Min Price", "Max Price",
               "Avg Rating", "Total Reviews", "Est. Sales", "Avg Stock", "Est. Revenue"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    brand_data = defaultdict(list)
    for r_d in cat_d:
        brand_data[r_d["marca"]].append(r_d)

    sorted_brands = sorted(brand_data.items(), key=lambda x: -sum(r["revenue_est"] for r in x[1]))
    
    for i, (brand, prods) in enumerate(sorted_brands, 1):
        row += 1
        prices = [r["preco_pix"] for r in prods if r["preco_pix"] > 0]
        ratings = [r["avaliacao"] for r in prods if r["avaliacao"] > 0]
        stocks = [r["estoque"] for r in prods]
        is_cm = brand.lower() == "cooler master"
        is_comp = brand in key_competitors
        
        write_row(ws, row, [
            i, brand, len(prods),
            statistics.mean(prices) if prices else 0,
            min(prices) if prices else 0,
            max(prices) if prices else 0,
            statistics.mean(ratings) if ratings else 0,
            sum(r["num_avaliacoes"] for r in prods),
            sum(r["vendas_estimadas"] for r in prods),
            statistics.mean(stocks) if stocks else 0,
            sum(r["revenue_est"] for r in prods),
        ], is_cm=is_cm)
        for c in [4,5,6,11]: fmt_brl(ws.cell(row=row, column=c))
        fmt_int(ws.cell(row=row, column=9))
        fmt_int(ws.cell(row=row, column=10))

    # CM vs Key Competitors detail
    row += 2
    write_section_title(ws, row, 1, f"CM vs Key Competitors — {cat_en} Detail", width=10)
    row += 1
    headers2 = ["Product", "Brand", "Price (PIX)", "Rating", "Reviews", "Est. Sales", "Stock", "Segment"]
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers2))

    focus = [r for r in cat_d if r["is_cm"] or r["marca"] in key_competitors]
    for p in sorted(focus, key=lambda x: -x["preco_pix"]):
        row += 1
        write_row(ws, row, [
            p["nome"], p["marca"], p["preco_pix"], p["avaliacao"],
            p["num_avaliacoes"], p["vendas_estimadas"], p["estoque"],
            price_segment(p["preco_pix"]),
        ], is_cm=p["is_cm"])
        fmt_brl(ws.cell(row=row, column=3))

    auto_width(ws)
    freeze(ws, "A4")
    add_filters(ws, 3, len(headers))
    print(f"  Competitive {cat_en}: rows up to {row}")


def build_brand_rankings(wb, data):
    ws = wb.create_sheet("Brand Rankings")
    ws.sheet_properties.tabColor = SUCCESS

    row = 1
    write_section_title(ws, row, 1, "Brand Rankings — Composite Score", width=12)

    row = 3
    headers = ["Rank", "Brand", "SKUs", "Categories", "Avg Price", "Avg Rating",
               "Total Reviews", "Est. Sales", "Est. Revenue", "Avg Stock",
               "Composite Score"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    brand_data = defaultdict(list)
    for r_d in data:
        brand_data[r_d["marca"]].append(r_d)

    # Compute composite: normalize each metric
    brand_metrics = {}
    for brand, prods in brand_data.items():
        prices = [r["preco_pix"] for r in prods if r["preco_pix"] > 0]
        ratings = [r["avaliacao"] for r in prods if r["avaliacao"] > 0]
        brand_metrics[brand] = {
            "skus": len(prods),
            "cats": len(set(r["categoria"] for r in prods)),
            "avg_price": statistics.mean(prices) if prices else 0,
            "avg_rating": statistics.mean(ratings) if ratings else 0,
            "total_reviews": sum(r["num_avaliacoes"] for r in prods),
            "est_sales": sum(r["vendas_estimadas"] for r in prods),
            "revenue": sum(r["revenue_est"] for r in prods),
            "avg_stock": statistics.mean([r["estoque"] for r in prods]),
        }

    # Simple composite: weighted normalized
    max_vals = {
        "skus": max(m["skus"] for m in brand_metrics.values()) or 1,
        "avg_rating": 5,
        "total_reviews": max(m["total_reviews"] for m in brand_metrics.values()) or 1,
        "revenue": max(m["revenue"] for m in brand_metrics.values()) or 1,
    }

    for brand, m in brand_metrics.items():
        m["score"] = (
            (m["skus"] / max_vals["skus"]) * 20 +
            (m["avg_rating"] / max_vals["avg_rating"]) * 30 +
            (m["total_reviews"] / max_vals["total_reviews"]) * 20 +
            (m["revenue"] / max_vals["revenue"]) * 30
        )

    sorted_brands = sorted(brand_metrics.items(), key=lambda x: -x[1]["score"])
    for i, (brand, m) in enumerate(sorted_brands, 1):
        row += 1
        is_cm = brand.lower() == "cooler master"
        write_row(ws, row, [
            i, brand, m["skus"], m["cats"], m["avg_price"], m["avg_rating"],
            m["total_reviews"], m["est_sales"], m["revenue"], m["avg_stock"],
            m["score"],
        ], is_cm=is_cm)
        fmt_brl(ws.cell(row=row, column=5))
        fmt_brl(ws.cell(row=row, column=9))
        fmt_int(ws.cell(row=row, column=8))

    auto_width(ws)
    freeze(ws, "A4")
    add_filters(ws, 3, len(headers))
    print(f"  Brand Rankings: rows up to {row}")


def build_ratings_reviews(wb, data):
    ws = wb.create_sheet("Ratings & Reviews")
    ws.sheet_properties.tabColor = SUCCESS

    row = 1
    write_section_title(ws, row, 1, "Top Products by Rating & Reviews", width=10)

    # Top 20 by rating (min 10 reviews)
    row = 3
    write_section_title(ws, row, 1, "Top 20 Highest Rated (min. 10 reviews)", width=8)
    row += 1
    headers = ["Rank", "Product", "Brand", "Category", "Rating", "Reviews", "Price (PIX)", "CM?"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    rated = [r for r in data if r["avaliacao"] > 0 and r["num_avaliacoes"] >= 10]
    for i, p in enumerate(sorted(rated, key=lambda x: (-x["avaliacao"], -x["num_avaliacoes"]))[:20], 1):
        row += 1
        write_row(ws, row, [i, p["nome"], p["marca"], p["cat_en"], p["avaliacao"],
                           p["num_avaliacoes"], p["preco_pix"], "✓" if p["is_cm"] else ""], is_cm=p["is_cm"])
        fmt_brl(ws.cell(row=row, column=7))

    # Top 20 by reviews
    row += 2
    write_section_title(ws, row, 1, "Top 20 Most Reviewed", width=8)
    row += 1
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    for i, p in enumerate(sorted(data, key=lambda x: -x["num_avaliacoes"])[:20], 1):
        row += 1
        write_row(ws, row, [i, p["nome"], p["marca"], p["cat_en"], p["avaliacao"],
                           p["num_avaliacoes"], p["preco_pix"], "✓" if p["is_cm"] else ""], is_cm=p["is_cm"])
        fmt_brl(ws.cell(row=row, column=7))

    # CM ranking
    row += 2
    write_section_title(ws, row, 1, "Cooler Master Rating Position", width=8)
    row += 1
    all_sorted = sorted(data, key=lambda x: (-x["avaliacao"], -x["num_avaliacoes"]))
    for i, p in enumerate(all_sorted, 1):
        if p["is_cm"]:
            row += 1
            write_row(ws, row, [f"#{i}/{len(all_sorted)}", p["nome"], p["marca"],
                               p["cat_en"], p["avaliacao"], p["num_avaliacoes"],
                               p["preco_pix"], "✓"], is_cm=True)
            fmt_brl(ws.cell(row=row, column=7))

    auto_width(ws)
    freeze(ws, "A5")
    print(f"  Ratings & Reviews: rows up to {row}")


def build_stock_availability(wb, data):
    ws = wb.create_sheet("Stock & Availability")
    ws.sheet_properties.tabColor = WARNING

    row = 1
    write_section_title(ws, row, 1, "Stock & Availability Analysis", width=10)

    row = 3
    headers = ["Category", "Total Products", "Available", "Unavailable", "Avg Stock",
               "Critical (<50)", "Low (50-200)", "Normal (200-2000)", "High (2000-5000)", "Mega (5000+)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    for cat_pt, cat_en in CAT_MAP.items():
        row += 1
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        avail = len([r for r in cat_d if r["disponivel"]])
        stocks = [r["estoque"] for r in cat_d]
        write_row(ws, row, [
            cat_en, len(cat_d), avail, len(cat_d) - avail,
            statistics.mean(stocks) if stocks else 0,
            len([s for s in stocks if s < 50]),
            len([s for s in stocks if 50 <= s < 200]),
            len([s for s in stocks if 200 <= s < 2000]),
            len([s for s in stocks if 2000 <= s < 5000]),
            len([s for s in stocks if s >= 5000]),
        ])
        fmt_int(ws.cell(row=row, column=5))

    # CM stock
    row += 2
    write_section_title(ws, row, 1, "Cooler Master Stock Detail", width=8)
    row += 1
    headers2 = ["Product", "Category", "Stock", "Status", "vs Cat. Avg", "Price"]
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers2))

    for p in [r for r in data if r["is_cm"]]:
        row += 1
        cat_avg_stock = statistics.mean([r["estoque"] for r in data if r["categoria"] == p["categoria"]])
        status = "Critical" if p["estoque"] < 50 else "Low" if p["estoque"] < 200 else "Normal" if p["estoque"] < 2000 else "High"
        write_row(ws, row, [
            p["nome"], p["cat_en"], p["estoque"], status,
            f"{(p['estoque'] / cat_avg_stock - 1) * 100:+.0f}%" if cat_avg_stock else "N/A",
            p["preco_pix"],
        ], is_cm=True)
        fmt_int(ws.cell(row=row, column=3))
        fmt_brl(ws.cell(row=row, column=6))
        # Color code status
        status_cell = ws.cell(row=row, column=4)
        if status == "Critical":
            status_cell.font = Font(bold=True, color=DANGER, size=10)
        elif status == "Low":
            status_cell.font = Font(bold=True, color=WARNING, size=10)
        elif status == "High":
            status_cell.font = Font(bold=True, color=SUCCESS, size=10)

    auto_width(ws)
    freeze(ws, "A4")
    print(f"  Stock & Availability: rows up to {row}")


def build_all_products(wb, data, cat_pt, cat_en):
    ws = wb.create_sheet(f"All Products — {cat_en}")
    ws.sheet_properties.tabColor = SUB_HEADER

    row = 1
    headers = ["#", "Product Name", "Brand", "SKU", "Price (PIX)", "List Price",
               "Discount %", "Rating", "Reviews", "Est. Sales", "Stock",
               "Segment", "Available", "Seller", "Subcategory", "CM?"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    cat_d = sorted([r for r in data if r["categoria"] == cat_pt], key=lambda x: -x["preco_pix"])
    for i, p in enumerate(cat_d, 1):
        row += 1
        write_row(ws, row, [
            i, p["nome"], p["marca"], p["partnumber"],
            p["preco_pix"], p["preco_lista"],
            p["desconto_pct"] / 100 if p["desconto_pct"] else 0,
            p["avaliacao"], p["num_avaliacoes"], p["vendas_estimadas"],
            p["estoque"], price_segment(p["preco_pix"]),
            "Yes" if p["disponivel"] else "No", p["seller_type"],
            p["subcategoria"], "✓" if p["is_cm"] else "",
        ], is_cm=p["is_cm"])
        fmt_brl(ws.cell(row=row, column=5))
        fmt_brl(ws.cell(row=row, column=6))
        fmt_pct(ws.cell(row=row, column=7))
        fmt_int(ws.cell(row=row, column=10))
        fmt_int(ws.cell(row=row, column=11))

    # Conditional formatting
    price_range = f"E2:E{row}"
    rating_range = f"H2:H{row}"
    stock_range = f"K2:K{row}"

    # Rating color scale
    ws.conditional_formatting.add(rating_range, ColorScaleRule(
        start_type="min", start_color="FEE2E2",
        mid_type="percentile", mid_value=50, mid_color="FEF3C7",
        end_type="max", end_color="D1FAE5",
    ))
    # Stock data bar
    ws.conditional_formatting.add(stock_range, DataBarRule(
        start_type="min", end_type="max", color=ACCENT_BLUE,
    ))

    auto_width(ws)
    freeze(ws, "A2")
    add_filters(ws, 1, len(headers))
    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    print(f"  All Products {cat_en}: {len(cat_d)} products, rows up to {row}")


def build_sales_analysis(wb, data):
    ws = wb.create_sheet("Sales Analysis")
    ws.sheet_properties.tabColor = "7C3AED"

    row = 1
    write_section_title(ws, row, 1, "Sales & Revenue Analysis", width=12)

    # Sales by category
    row = 3
    headers = ["Category", "Total Est. Sales", "Total Est. Revenue", "Avg Revenue/SKU", "Top Seller"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    for cat_pt, cat_en in CAT_MAP.items():
        row += 1
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        total_sales = sum(r["vendas_estimadas"] for r in cat_d)
        total_rev = sum(r["revenue_est"] for r in cat_d)
        top = max(cat_d, key=lambda x: x["vendas_estimadas"])
        write_row(ws, row, [cat_en, total_sales, total_rev, total_rev / len(cat_d) if cat_d else 0,
                           f"{top['marca']} - {top['nome'][:40]}"])
        fmt_int(ws.cell(row=row, column=2))
        fmt_brl(ws.cell(row=row, column=3))
        fmt_brl(ws.cell(row=row, column=4))

    # Top 30 by sales
    row += 2
    write_section_title(ws, row, 1, "Top 30 Best-Selling Products (by Est. Sales)", width=10)
    row += 1
    headers2 = ["Rank", "Product", "Brand", "Category", "Price (PIX)", "Est. Sales",
                "Est. Revenue", "Rating", "Reviews", "Stock", "CM?"]
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers2))

    for i, p in enumerate(sorted(data, key=lambda x: -x["vendas_estimadas"])[:30], 1):
        row += 1
        write_row(ws, row, [
            i, p["nome"], p["marca"], p["cat_en"], p["preco_pix"],
            p["vendas_estimadas"], p["revenue_est"], p["avaliacao"],
            p["num_avaliacoes"], p["estoque"], "✓" if p["is_cm"] else "",
        ], is_cm=p["is_cm"])
        fmt_brl(ws.cell(row=row, column=5))
        fmt_int(ws.cell(row=row, column=6))
        fmt_brl(ws.cell(row=row, column=7))

    # Revenue by brand
    row += 2
    write_section_title(ws, row, 1, "Revenue by Brand — Market Share", width=10)
    row += 1
    headers3 = ["Rank", "Brand", "Est. Revenue", "Revenue Share %", "Est. Sales", "SKUs", "Avg Price"]
    for c, h in enumerate(headers3, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers3))

    brand_rev = defaultdict(float)
    brand_sales = defaultdict(int)
    brand_skus = defaultdict(int)
    brand_prices = defaultdict(list)
    for r_d in data:
        brand_rev[r_d["marca"]] += r_d["revenue_est"]
        brand_sales[r_d["marca"]] += r_d["vendas_estimadas"]
        brand_skus[r_d["marca"]] += 1
        if r_d["preco_pix"] > 0:
            brand_prices[r_d["marca"]].append(r_d["preco_pix"])

    total_rev = sum(brand_rev.values())
    for i, (brand, rev) in enumerate(sorted(brand_rev.items(), key=lambda x: -x[1]), 1):
        row += 1
        is_cm = brand.lower() == "cooler master"
        avg_p = statistics.mean(brand_prices[brand]) if brand_prices[brand] else 0
        write_row(ws, row, [
            i, brand, rev, rev / total_rev if total_rev else 0,
            brand_sales[brand], brand_skus[brand], avg_p,
        ], is_cm=is_cm)
        fmt_brl(ws.cell(row=row, column=3))
        fmt_pct(ws.cell(row=row, column=4))
        fmt_int(ws.cell(row=row, column=5))
        fmt_brl(ws.cell(row=row, column=7))
        if i > 30:
            break

    # CM sales ranking by category
    row += 2
    write_section_title(ws, row, 1, "CM Sales Position by Category", width=8)
    row += 1
    for c, h in enumerate(["Category", "CM Sales Rank", "CM Revenue Rank", "CM Revenue", "Category Total Rev", "CM Share"], 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 6)

    for cat_pt, cat_en in CAT_MAP.items():
        row += 1
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        brand_cat_rev = defaultdict(float)
        brand_cat_sales = defaultdict(int)
        for r_d in cat_d:
            brand_cat_rev[r_d["marca"]] += r_d["revenue_est"]
            brand_cat_sales[r_d["marca"]] += r_d["vendas_estimadas"]
        
        rev_rank = sorted(brand_cat_rev, key=lambda b: -brand_cat_rev[b])
        sales_rank = sorted(brand_cat_sales, key=lambda b: -brand_cat_sales[b])
        cm_rev_pos = (rev_rank.index("Cooler Master") + 1) if "Cooler Master" in rev_rank else "N/A"
        cm_sales_pos = (sales_rank.index("Cooler Master") + 1) if "Cooler Master" in sales_rank else "N/A"
        cm_rev = brand_cat_rev.get("Cooler Master", 0)
        cat_total = sum(brand_cat_rev.values())
        
        write_row(ws, row, [
            cat_en,
            f"#{cm_sales_pos}/{len(sales_rank)}" if isinstance(cm_sales_pos, int) else "N/A",
            f"#{cm_rev_pos}/{len(rev_rank)}" if isinstance(cm_rev_pos, int) else "N/A",
            cm_rev, cat_total,
            cm_rev / cat_total if cat_total else 0,
        ], is_cm=True)
        fmt_brl(ws.cell(row=row, column=4))
        fmt_brl(ws.cell(row=row, column=5))
        fmt_pct(ws.cell(row=row, column=6))

    auto_width(ws)
    freeze(ws, "A4")
    print(f"  Sales Analysis: rows up to {row}")


def build_revenue_estimation(wb, data):
    ws = wb.create_sheet("Revenue Estimation")
    ws.sheet_properties.tabColor = "059669"

    row = 1
    write_section_title(ws, row, 1, "Revenue Estimation — Full Breakdown", width=12)
    row += 1
    note = ws.cell(row=row, column=1, value="Note: Revenue = Price (PIX) × Est. Sales (num_avaliacoes × 7). Approximation based on review velocity.")
    note.font = font_small
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # Revenue by category
    row += 2
    headers = ["Category", "Total Revenue", "% of Grand Total", "Total Sales Units", "Avg Revenue/Product", "# Products"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    grand_total = sum(r["revenue_est"] for r in data)
    for cat_pt, cat_en in CAT_MAP.items():
        row += 1
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        rev = sum(r["revenue_est"] for r in cat_d)
        sales = sum(r["vendas_estimadas"] for r in cat_d)
        write_row(ws, row, [cat_en, rev, rev / grand_total if grand_total else 0,
                           sales, rev / len(cat_d) if cat_d else 0, len(cat_d)])
        fmt_brl(ws.cell(row=row, column=2))
        fmt_pct(ws.cell(row=row, column=3))
        fmt_int(ws.cell(row=row, column=4))
        fmt_brl(ws.cell(row=row, column=5))

    # Grand total
    row += 1
    total_sales = sum(r["vendas_estimadas"] for r in data)
    write_row(ws, row, ["TOTAL", grand_total, 1.0, total_sales, grand_total / len(data), len(data)], bold=True)
    fmt_brl(ws.cell(row=row, column=2))
    fmt_pct(ws.cell(row=row, column=3))
    fmt_int(ws.cell(row=row, column=4))
    fmt_brl(ws.cell(row=row, column=5))

    # CM vs Top 10
    row += 2
    write_section_title(ws, row, 1, "CM Revenue vs Top 10 Competitors", width=10)
    row += 1
    headers2 = ["Rank", "Brand", "Est. Revenue", "Revenue Share", "# Products", "Avg Price", "Total Sales"]
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers2))

    brand_rev = defaultdict(float)
    brand_sales = defaultdict(int)
    brand_cnt = defaultdict(int)
    brand_prices = defaultdict(list)
    for r_d in data:
        brand_rev[r_d["marca"]] += r_d["revenue_est"]
        brand_sales[r_d["marca"]] += r_d["vendas_estimadas"]
        brand_cnt[r_d["marca"]] += 1
        if r_d["preco_pix"] > 0:
            brand_prices[r_d["marca"]].append(r_d["preco_pix"])

    for i, (brand, rev) in enumerate(sorted(brand_rev.items(), key=lambda x: -x[1])[:15], 1):
        row += 1
        is_cm = brand.lower() == "cooler master"
        write_row(ws, row, [
            i, brand, rev, rev / grand_total,
            brand_cnt[brand],
            statistics.mean(brand_prices[brand]) if brand_prices[brand] else 0,
            brand_sales[brand],
        ], is_cm=is_cm)
        fmt_brl(ws.cell(row=row, column=3))
        fmt_pct(ws.cell(row=row, column=4))
        fmt_brl(ws.cell(row=row, column=6))
        fmt_int(ws.cell(row=row, column=7))

    # Revenue by segment
    row += 2
    write_section_title(ws, row, 1, "Revenue by Price Segment", width=8)
    row += 1
    headers3 = ["Segment", "Total Revenue", "Revenue Share", "# Products", "Avg Revenue/Product", "CM Revenue", "CM Share"]
    for c, h in enumerate(headers3, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers3))

    for seg in SEGMENTS_ORDER:
        row += 1
        seg_d = [r for r in data if price_segment(r["preco_pix"]) == seg]
        rev = sum(r["revenue_est"] for r in seg_d)
        cm_rev = sum(r["revenue_est"] for r in seg_d if r["is_cm"])
        write_row(ws, row, [
            seg, rev, rev / grand_total if grand_total else 0,
            len(seg_d), rev / len(seg_d) if seg_d else 0,
            cm_rev, cm_rev / rev if rev else 0,
        ], is_cm=cm_rev > 0)
        fmt_brl(ws.cell(row=row, column=2))
        fmt_pct(ws.cell(row=row, column=3))
        fmt_brl(ws.cell(row=row, column=5))
        fmt_brl(ws.cell(row=row, column=6))
        fmt_pct(ws.cell(row=row, column=7))

    # Revenue by category per brand (top 5 per category)
    row += 2
    write_section_title(ws, row, 1, "Revenue Share by Category — Top Brands", width=8)
    for cat_pt, cat_en in CAT_MAP.items():
        row += 1
        ws.cell(row=row, column=1, value=cat_en).font = font_bold
        row += 1
        for c, h in enumerate(["Rank", "Brand", "Revenue", "Share %", "SKUs"], 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, 5)
        
        cat_d = [r for r in data if r["categoria"] == cat_pt]
        cat_brand_rev = defaultdict(float)
        cat_brand_cnt = defaultdict(int)
        for r_d in cat_d:
            cat_brand_rev[r_d["marca"]] += r_d["revenue_est"]
            cat_brand_cnt[r_d["marca"]] += 1
        cat_total = sum(cat_brand_rev.values())
        
        for i, (brand, rev) in enumerate(sorted(cat_brand_rev.items(), key=lambda x: -x[1])[:8], 1):
            row += 1
            is_cm = brand.lower() == "cooler master"
            write_row(ws, row, [i, brand, rev, rev / cat_total if cat_total else 0, cat_brand_cnt[brand]], is_cm=is_cm)
            fmt_brl(ws.cell(row=row, column=3))
            fmt_pct(ws.cell(row=row, column=4))

    auto_width(ws)
    freeze(ws, "A5")
    print(f"  Revenue Estimation: rows up to {row}")


def build_trend_tab(wb, data, period_name, note_text):
    ws = wb.create_sheet(period_name)
    ws.sheet_properties.tabColor = "6366F1"

    row = 1
    write_section_title(ws, row, 1, f"{period_name} — Historical Tracking", width=12)

    row = 2
    n = ws.cell(row=row, column=1, value=note_text)
    n.font = Font(italic=True, color="6B7280", size=10, name="Calibri")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # Structure with current snapshot
    row = 4
    write_section_title(ws, row, 1, "Current Snapshot — Price & Stock Baseline", width=10)
    row += 1
    headers = ["Product", "Brand", "Category", "Current Price", "List Price",
               "Discount %", "Stock", "Rating", "Est. Sales", "Data Date"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    # Show CM products + top sellers as baseline
    focus = sorted([r for r in data if r["is_cm"]], key=lambda x: -x["preco_pix"])
    focus += sorted([r for r in data if not r["is_cm"]], key=lambda x: -x["vendas_estimadas"])[:20]
    
    for p in focus:
        row += 1
        write_row(ws, row, [
            p["nome"], p["marca"], p["cat_en"], p["preco_pix"], p["preco_lista"],
            p["desconto_pct"] / 100 if p["desconto_pct"] else 0,
            p["estoque"], p["avaliacao"], p["vendas_estimadas"], p["data_coleta"][:10],
        ], is_cm=p["is_cm"])
        fmt_brl(ws.cell(row=row, column=4))
        fmt_brl(ws.cell(row=row, column=5))
        fmt_pct(ws.cell(row=row, column=6))
        fmt_int(ws.cell(row=row, column=7))

    # Placeholder sections
    row += 2
    write_section_title(ws, row, 1, "Price Variation (Δ%)", width=6)
    row += 1
    for c, h in enumerate(["Product", "Brand", "Start Price", "End Price", "Change %", "Direction"], 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 6)
    row += 1
    ws.cell(row=row, column=1, value="⏳ Awaiting historical data from daily scraper...").font = font_small
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    row += 2
    write_section_title(ws, row, 1, "New Products Added", width=4)
    row += 1
    for c, h in enumerate(["Product", "Brand", "Price", "Date Added"], 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 4)
    row += 1
    ws.cell(row=row, column=1, value="⏳ Will track new listings automatically...").font = font_small

    row += 2
    write_section_title(ws, row, 1, "Stock Changes", width=4)
    row += 1
    for c, h in enumerate(["Product", "Brand", "Previous Stock", "Current Stock"], 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 4)
    row += 1
    ws.cell(row=row, column=1, value="⏳ Will populate with next scrape cycle...").font = font_small

    auto_width(ws)
    freeze(ws, "A6")
    print(f"  {period_name}: rows up to {row}")


def build_data_dictionary(wb):
    ws = wb.create_sheet("Data Dictionary")
    ws.sheet_properties.tabColor = "9CA3AF"

    row = 1
    write_section_title(ws, row, 1, "Data Dictionary & Methodology", width=6)

    entries = [
        ("", ""),
        ("Source", "KaBuM.com.br — Brazil's largest hardware e-commerce platform"),
        ("Collection Date", "February 16, 2026"),
        ("Collection Method", "Automated web scraper (Python/Playwright)"),
        ("Categories", "PSU (Fontes), Cases (Gabinetes), Coolers"),
        ("Total Products", "403"),
        ("", ""),
        ("FIELD DEFINITIONS", ""),
        ("preco_brl", "Best available price in BRL"),
        ("preco_pix", "Price for PIX payment (instant bank transfer, usually best price)"),
        ("preco_lista", "Listed/MSRP price before promotions"),
        ("desconto_pct", "Discount percentage from list price"),
        ("vendas_estimadas", "Estimated total sales = num_avaliacoes × 7 (review-to-sales ratio)"),
        ("estoque", "Current stock level reported by KaBuM"),
        ("avaliacao", "Product rating (0-5 stars)"),
        ("num_avaliacoes", "Number of customer reviews"),
        ("revenue_est", "Estimated revenue = preco_pix × vendas_estimadas"),
        ("is_cooler_master", "Product identified as Cooler Master brand"),
        ("", ""),
        ("PRICE SEGMENTS", ""),
        ("Budget", "< R$200"),
        ("Mid", "R$200 — R$500"),
        ("Premium", "R$500 — R$1,000"),
        ("Ultra-Premium", "> R$1,000"),
        ("", ""),
        ("METHODOLOGY NOTES", ""),
        ("Sales Estimation", "Based on review count × 7 multiplier. Industry standard approximation."),
        ("Revenue", "Gross estimated revenue. Does not account for returns, marketplace fees."),
        ("Market Share", "Based on SKU count or estimated revenue, as labeled."),
        ("Composite Score", "Weighted: Rating (30%) + Revenue (30%) + SKUs (20%) + Reviews (20%)"),
        ("", ""),
        ("DISCLAIMER", ""),
        ("", "This report is for internal use only. Data is approximate and based on publicly"),
        ("", "available information from KaBuM.com.br. Actual sales figures may differ."),
        ("", "Prepared by Cooler Master Brazil — Country Manager: Limudo"),
        ("", "For questions: contact LATAM Manager Ariel Mai"),
        ("", ""),
        ("Q1 2026 Context", "Brasil confirmed USD $3,132,691.77, 91% hit rate"),
    ]

    for label, desc in entries:
        row += 1
        if label and not desc:  # Section header
            ws.cell(row=row, column=1, value=label).font = font_bold
            ws.cell(row=row, column=1).fill = fill_light_blue
            ws.cell(row=row, column=2).fill = fill_light_blue
        elif label:
            ws.cell(row=row, column=1, value=label).font = font_bold
            ws.cell(row=row, column=2, value=desc).font = font_normal
        else:
            ws.cell(row=row, column=2, value=desc).font = font_normal

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80
    freeze(ws, "A2")
    print(f"  Data Dictionary: rows up to {row}")


# ═══════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════

def main():
    print("Loading data...")
    data = load_data()
    print(f"  Loaded {len(data)} products")
    print(f"  Categories: { {CAT_MAP.get(c,c): len([r for r in data if r['categoria']==c]) for c in set(r['categoria'] for r in data)} }")
    print(f"  CM products: {len([r for r in data if r['is_cm']])}")

    wb = Workbook()
    
    print("\nBuilding tabs...")
    build_executive_summary(wb, data)
    build_market_overview(wb, data)
    build_cm_analysis(wb, data)
    build_price_positioning(wb, data)
    
    # Competitive tabs
    build_competitive_tab(wb, data, "fontes", "PSU",
        ["Corsair", "EVGA", "Seasonic", "MSI", "NZXT", "DeepCool"])
    build_competitive_tab(wb, data, "gabinetes", "Cases",
        ["Corsair", "NZXT", "DeepCool", "Lian Li", "MSI", "Montech"])
    build_competitive_tab(wb, data, "coolers", "Coolers",
        ["Noctua", "be quiet!", "DeepCool", "Corsair", "Thermalright"])
    
    build_brand_rankings(wb, data)
    build_ratings_reviews(wb, data)
    build_stock_availability(wb, data)
    build_sales_analysis(wb, data)
    build_revenue_estimation(wb, data)
    
    # Trend tabs
    build_trend_tab(wb, data, "7-Day Trend",
        "⚠️ Historical data will populate automatically via daily scraper. Currently showing Day 1 baseline (2026-02-16).")
    build_trend_tab(wb, data, "Monthly Summary",
        "⚠️ Monthly trends will build as the scraper collects daily snapshots through February 2026.")
    build_trend_tab(wb, data, "Quarterly (Q1 2026)",
        "⚠️ Q1 2026 (Jan–Mar) tracking. Brasil target: USD $3,132,691.77 (91% hit rate). Daily data accumulates automatically.")
    
    # All products tabs
    for cat_pt, cat_en in CAT_MAP.items():
        build_all_products(wb, data, cat_pt, cat_en)
    
    build_data_dictionary(wb)

    # Save
    print(f"\nTotal sheets: {len(wb.sheetnames)}")
    print(f"Sheets: {wb.sheetnames}")
    
    primary = OUT_PATHS[0]
    os.makedirs(os.path.dirname(primary), exist_ok=True)
    wb.save(primary)
    size = os.path.getsize(primary)
    print(f"\nSaved primary: {primary} ({size/1024:.0f} KB)")

    for p in OUT_PATHS[1:]:
        try:
            os.makedirs(os.path.dirname(p), exist_ok=True)
            shutil.copy2(primary, p)
            print(f"Copied to: {p}")
        except Exception as e:
            print(f"Could not copy to {p}: {e}")

    print("\n✅ Dashboard complete!")


if __name__ == "__main__":
    main()
