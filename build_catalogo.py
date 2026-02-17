#!/usr/bin/env python3
"""
CM_Catalogo_Brasil — Pipeline completo
- Carrega Packing List CM (XLS) + Scrape KaBuM (CSV)
- Cruza Part Numbers (exact + fuzzy)
- Busca links oficiais dos fabricantes (sitemap CM + padrões de URL)
- Gera Excel multi-abas + Dashboard HTML
"""

import csv, re, json, ssl, time, gzip, sys, os
from pathlib import Path
from collections import defaultdict, Counter
from urllib.request import urlopen, Request
from urllib.error import HTTPError, URLError
from datetime import datetime
import xlrd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from thefuzz import fuzz, process as fuzz_process

# ─── Paths ───────────────────────────────────────────────────────────────────
PACKING_LIST = Path.home() / "Library/CloudStorage/OneDrive-CoolerMaster/2026/PackingList_APAC_MALA.xls"
KABUM_DIR    = Path.home() / ".openclaw/workspace-pesquisa/kabum-scraper/data/historico"
OUTPUT_DIR   = Path.home() / ".openclaw/workspace-main"
ONEDRIVE_DIR = Path.home() / "Library/CloudStorage/OneDrive-CoolerMaster/2026/Dashboards"

EXCEL_OUT = OUTPUT_DIR / "CM_Catalogo_Brasil.xlsx"
HTML_OUT  = OUTPUT_DIR / "CM_Catalogo_Dashboard.html"

SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS_HTTP = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

# ─── Colors ──────────────────────────────────────────────────────────────────
COL_HEADER_DARK  = "1A2B4A"   # azul escuro
COL_HEADER_TEXT  = "FFFFFF"
COL_ACTIVE       = "E8F5E9"
COL_EOL          = "FFEBEE"
COL_ROW_ALT      = "F5F7FA"
COL_GREEN        = "27AE60"
COL_RED          = "E74C3C"
COL_ORANGE       = "F39C12"

# ─── 1. Load KaBuM CSV ───────────────────────────────────────────────────────
def load_kabum():
    """Carrega o CSV mais recente do scrape KaBuM"""
    csvs = sorted(KABUM_DIR.glob("kabum_*_full_clean.csv"))
    if not csvs:
        csvs = sorted(KABUM_DIR.glob("kabum_*.csv"))
    latest = csvs[-1]
    print(f"  KaBuM CSV: {latest.name}")
    rows = []
    with open(latest, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    print(f"  Total produtos KaBuM: {len(rows)}")
    return rows

# ─── 2. Load Packing List ────────────────────────────────────────────────────
def load_packing_list():
    """Carrega o Packing List CM (XLS)"""
    wb = xlrd.open_workbook(str(PACKING_LIST))
    sh = wb.sheet_by_index(0)
    headers = [sh.cell_value(0, c) for c in range(sh.ncols)]
    rows = []
    for r in range(1, sh.nrows):
        row = {}
        for c, h in enumerate(headers):
            v = sh.cell_value(r, c)
            row[h] = v
        rows.append(row)
    print(f"  Packing List: {len(rows)} produtos, {len(headers)} colunas")
    return rows

# ─── 3. Cooler Master Sitemap ─────────────────────────────────────────────────
def _fetch_cm_sitemap(url: str) -> dict:
    """Baixa um sitemap CM e retorna slug→url map"""
    try:
        req = Request(url, headers={**HEADERS_HTTP, "Accept-Encoding": "gzip"})
        resp = urlopen(req, context=SSL_CTX, timeout=30)
        data = resp.read()
        if resp.headers.get("Content-Encoding") == "gzip":
            data = gzip.decompress(data)
        text = data.decode("utf-8", errors="ignore")
        urls = re.findall(r"<loc>(https://www\.coolermaster\.com[^<]+)</loc>", text)
        slug_map = {}
        for u in urls:
            m = re.search(r"/products/([^/]+)/?$", u)
            if m:
                slug = m.group(1).lower()
                # Keep pt-br preferred, but don't overwrite
                if slug not in slug_map:
                    slug_map[slug] = u
        return slug_map
    except Exception as e:
        print(f"  WARN: Não foi possível baixar {url}: {e}")
        return {}

def load_cm_sitemap():
    """Baixa sitemaps PT-BR + EN-Global da CM, combina em um mapa único"""
    print(f"  Baixando sitemaps CM (pt-br + en-global)...")
    # pt-br tem prioridade, en-global é fallback
    ptbr_map = _fetch_cm_sitemap("https://www.coolermaster.com/pt-br-sitemap-products.xml")
    englobal_map = _fetch_cm_sitemap("https://www.coolermaster.com/en-global-sitemap-products.xml")
    
    # Merge: pt-br preferred, en-global fills gaps
    combined = {**englobal_map, **ptbr_map}  # pt-br overwrites en-global
    print(f"  CM sitemap: pt-br={len(ptbr_map)}, en-global={len(englobal_map)}, combinado={len(combined)} slugs")
    return combined

# ─── 4. Match CM products to sitemap ─────────────────────────────────────────
def _extract_model_keywords(nome: str) -> list:
    """Extrai keywords do modelo a partir do nome do produto KaBuM"""
    # Remove prefixos de categoria e marca
    stopwords = {"fonte", "gabinete", "gamer", "air", "cooler", "master", "pasta",
                 "térmica", "ventoinha", "kit", "processador", "com", "sem", "cabo",
                 "pfc", "ativo", "atx", "full", "modular", "preto", "branco", "branca",
                 "plus", "series", "edition", "mid", "tower", "lateral", "vidro",
                 "temperado", "rgb", "argb", "led", "intel", "amd", "fan", "fans",
                 "flow", "mesh", "frontal", "black", "white", "wifi", "watt"}
    
    # Pega só o nome antes do primeiro '-' que separa SKU, ou antes da última spec
    clean = nome
    # Remove o suffix de SKU (após ' - ' no final)
    clean = re.sub(r'\s+-\s+[A-Z0-9\-]+\s*$', '', clean)
    # Remove tudo após vírgula (specs numéricas)
    parts = clean.split(',')
    clean = parts[0]
    
    # Lowercase e tokenize
    tokens = re.findall(r'[a-zA-Z0-9]+', clean.lower())
    
    # Remove stopwords e tokens muito curtos
    # Permite tokens de 2 chars se forem alfanuméricos e contiverem dígito (v2, v3, q5, etc.)
    def keep(t):
        if t in stopwords: return False
        if len(t) >= 3: return True
        if len(t) == 2 and re.match(r'^[a-z]\d$', t): return True  # v2, v3, q5
        return False
    keywords = [t for t in tokens if keep(t)]
    return keywords

def _score_slug(slug: str, keywords: list) -> float:
    """Pontua um slug com base em quantas keywords ele contém"""
    slug_lower = slug.lower()
    if not keywords:
        return 0
    # Cada keyword presente no slug vale 1 ponto
    hits = sum(1 for kw in keywords if kw in slug_lower)
    # Bonus: todos os keywords presentes
    all_hit_bonus = 3 if hits == len(keywords) else 0
    # Penalidade suave por slug muito mais longo que as keywords
    slug_tokens = slug.split('-')
    length_penalty = max(0, (len(slug_tokens) - len(keywords) - 3) * 0.05)
    return hits + all_hit_bonus - length_penalty

# Manual overrides for specific part numbers that are hard to match automatically
CM_PN_OVERRIDES = {
    # Thermal pastes
    "HTK-002-U1":     "https://www.coolermaster.com/pt-br/products/high-performance-thermal-paste/",
    "HTK-002-U1-GP":  "https://www.coolermaster.com/pt-br/products/high-performance-thermal-paste/",
    "MGZ-NDBG-N40G-R1": "https://www.coolermaster.com/pt-br/products/mastergel-maker/",
    "MGY-NOSG-N07M-R1": "https://www.coolermaster.com/pt-br/products/cryofuze-5/",
    # Cases — MCB-D500D = TD500 Mesh (D = mesh/argb variant)
    "MCB-D500D-WGNN-S00": "https://www.coolermaster.com/pt-br/products/masterbox-td500-argb/",
    "MCB-D500D-WGNN-S01": "https://www.coolermaster.com/pt-br/products/masterbox-td500-argb/",
    "MCB-D500D-WGNN-SER": "https://www.coolermaster.com/pt-br/products/masterbox-td500-argb/",
    "MCB-D500D-WGNN-SPS": "https://www.coolermaster.com/pt-br/products/masterbox-td500-argb/",
    "MCB-D500D-WGNN-STU": "https://www.coolermaster.com/pt-br/products/masterbox-td500-argb/",
}

def find_cm_url(nome: str, part_no: str, slug_map: dict, item_desc: str = "") -> str:
    """
    Tenta achar URL oficial CM via sitemap usando keyword matching.
    Combina nome do KaBuM e Item_Desc do Packing List.
    """
    # 0. Check manual overrides first
    if part_no and part_no.upper() in CM_PN_OVERRIDES:
        return CM_PN_OVERRIDES[part_no.upper()]
    
    if not slug_map:
        return ""
    
    # Extrai keywords de ambas as fontes
    keywords_kabum = _extract_model_keywords(nome)
    keywords_pl    = _extract_model_keywords(item_desc) if item_desc else []
    
    # Une keywords únicas, priorizando as do nome KaBuM
    all_kws = list(dict.fromkeys(keywords_kabum + keywords_pl))
    
    if not all_kws:
        return ""
    
    # Scoring: encontra o melhor slug
    # Em caso de empate: prefere slug mais curto (match mais preciso) e pt-br
    best_slug = None
    best_score = 0
    
    for slug in slug_map:
        score = _score_slug(slug, all_kws)
        if score > best_score + 0.01:  # clear winner
            best_score = score
            best_slug = slug
        elif abs(score - best_score) <= 0.01 and best_slug:
            # Tiebreaker: prefer shorter slug (more precise match)
            if len(slug) < len(best_slug):
                best_slug = slug
    
    # Threshold mínimo: score >= 2.5 (garante pelo menos 2 keywords + parcial)
    if best_score >= 2.5 and best_slug:
        return slug_map[best_slug]
    
    return ""

# ─── 5. Other Brand URL lookup ────────────────────────────────────────────────
BRAND_URL_PATTERNS = {
    "Corsair": [
        "https://www.corsair.com/br/pt/p/{pn_lower}",
        "https://www.corsair.com/br/pt/Categories/{category}---/p/{pn}",
    ],
    "NZXT": [
        "https://nzxt.com/pt-BR/product/{slug}",
        "https://nzxt.com/product/{slug}",
    ],
    "Deepcool": [
        "https://www.deepcool.com/products/{slug}",
        "https://www.deepcool.com/product/{pn_lower}",
    ],
    "MSI": [
        "https://www.msi.com/power-supply/{slug}",
        "https://www.msi.com/pc/components/{pn}",
    ],
    "ASUS": [
        "https://www.asus.com/br/motherboards-components/power-supply-units/{slug}/",
        "https://www.asus.com/br/accessories/{slug}/",
    ],
    "Lian Li": [
        "https://lian-li.com/product/{slug}/",
    ],
    "MONTECH": [
        "https://www.montech.com/product/{slug}/",
        "https://www.montech.com.tw/product/{slug}/",
    ],
    "Redragon": [
        "https://www.redragonbr.com.br/produto/{slug}",
    ],
    "Rise Mode": [
        "https://www.risemode.com.br/produto/{slug}",
    ],
    "Gigabyte": [
        "https://www.gigabyte.com/Power-Supply/{slug}",
    ],
    "Husky": [
        "https://www.huskytech.com.br/produto/{slug}",
        "https://www.husky.com.br/produtos/{slug}",
    ],
}

def _extract_model_slug(nome: str, brand: str) -> str:
    """Extrai slug do modelo a partir do nome do produto"""
    # Remove prefixo do brand e categoria
    clean = nome
    for prefix in [brand, "Gabinete", "Fonte", "Gamer", "Air Cooler", "Cooler", "Ventoinha", "Pasta Térmica"]:
        clean = re.sub(r'\b' + re.escape(prefix) + r'\b', '', clean, flags=re.IGNORECASE)
    # Remove tudo após vírgula (specs)
    clean = re.sub(r',.*', '', clean)
    clean = clean.strip()
    # Build slug
    slug = re.sub(r'[^a-zA-Z0-9\-\s]', '', clean).strip()
    slug = re.sub(r'[\s]+', '-', slug).lower().strip('-')
    return slug

def head_request(url: str, timeout: int = 5) -> bool:
    """Retorna True se URL responde 200 ou 3xx"""
    try:
        req = Request(url, method="HEAD", headers=HEADERS_HTTP)
        resp = urlopen(req, context=SSL_CTX, timeout=timeout)
        return resp.status < 400
    except HTTPError as e:
        return e.code < 400
    except Exception:
        return False

def find_brand_url(nome: str, part_no: str, brand: str, fab_url: str) -> str:
    """Tenta encontrar URL oficial do fabricante para produto não-CM"""
    if not brand:
        return ""
    
    slug = _extract_model_slug(nome, brand)
    pn_lower = part_no.lower().replace(' ', '-') if part_no else ''
    pn = part_no or ''
    
    patterns = BRAND_URL_PATTERNS.get(brand, [])
    
    for pattern in patterns:
        try:
            url = pattern.format(slug=slug, pn=pn, pn_lower=pn_lower, category="power-supplies")
            if head_request(url):
                return url
        except Exception:
            pass
    
    # Fallback: fabricante_url from scrape data
    return fab_url or ""

# ─── 6. Build manufacturer URL map (com cache) ───────────────────────────────
def build_url_map(kabum_rows: list, cm_slug_map: dict) -> dict:
    """
    Para cada produto KaBuM, tenta encontrar URL oficial.
    Retorna dict: product_key → official_url
    """
    cache_path = OUTPUT_DIR / ".url_cache.json"
    cache = {}
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text())
            print(f"  Cache de URLs: {len(cache)} entradas")
        except Exception:
            pass
    
    result = {}
    total = len(kabum_rows)
    cm_found = 0
    other_found = 0
    
    # Build PL item_desc lookup by KaBuM id (for CM products)
    # We do this so we can pass the official Item_Desc to find_cm_url
    # Note: pl_kb_matches not available here, so we'll do a quick direct match
    kb_pn_to_item_desc = {}
    
    for i, row in enumerate(kabum_rows):
        nome     = row.get("nome", "")
        marca    = row.get("marca", "")
        pn       = row.get("partnumber", "")
        fab_url  = row.get("fabricante_url", "")
        is_cm    = row.get("is_cooler_master", "").strip().lower() == "true"
        key      = row.get("id", str(i))
        
        if key in cache:
            result[key] = cache[key]
            continue
        
        if is_cm or marca.lower() == "cooler master":
            item_desc = kb_pn_to_item_desc.get(pn, "")
            url = find_cm_url(nome, pn, cm_slug_map, item_desc)
            if url:
                cm_found += 1
            result[key] = url
        else:
            # Para outras marcas: tenta padrões conhecidos
            url = find_brand_url(nome, pn, marca, fab_url)
            if url and url.startswith("http") and not url == fab_url:
                other_found += 1
            # Se não achou URL específica, mantém a base do fabricante
            result[key] = url if url else fab_url
        
        cache[key] = result[key]
        
        if (i + 1) % 50 == 0:
            print(f"    URLs: {i+1}/{total} processados...")
            cache_path.write_text(json.dumps(cache, indent=2))
    
    cache_path.write_text(json.dumps(cache, indent=2))
    print(f"  CM URLs encontradas: {cm_found} | Outras marcas: {other_found}")
    return result

# ─── 7. Cross-reference Packing List × KaBuM ─────────────────────────────────
def cross_reference(pl_rows: list, kabum_rows: list) -> dict:
    """
    Cruza Part Numbers do Packing List com SKUs da KaBuM.
    Retorna dict: part_no_CM → kabum_row (ou None)
    """
    # Build KaBuM index
    kb_by_pn  = {}  # exact match por partnumber
    kb_by_name = []  # para fuzzy
    
    for row in kabum_rows:
        pn = row.get("partnumber", "").strip().upper()
        if pn:
            kb_by_pn[pn] = row
        kb_by_name.append((row.get("nome", "").lower(), row))
    
    matches = {}
    fuzzy_matches = {}
    
    for pl_row in pl_rows:
        pn = str(pl_row.get("Part No", "")).strip().upper()
        if not pn:
            continue
        
        # 1. Exact match
        if pn in kb_by_pn:
            matches[pn] = kb_by_pn[pn]
            continue
        
        # 2. Partial match (sem sufixo de região) — requer base ≥ 8 chars
        # Remove sufixo regional (últimos 4 chars se forem letras como -BBR, -WW, -EU)
        if len(pn) > 4 and re.match(r'^[A-Z]{2,4}$', pn[-3:]):
            base_pn = pn[:-4]  # remove -BBR, -WW, etc.
        elif len(pn) > 5 and re.match(r'^[A-Z]{2,4}$', pn[-4:]):
            base_pn = pn[:-5]
        else:
            base_pn = pn
        
        # Só faz partial match se a base for longa o suficiente (evita CH, SH, etc.)
        if len(base_pn) >= 8:
            for kb_pn, kb_row in kb_by_pn.items():
                if kb_pn.startswith(base_pn) or base_pn in kb_pn:
                    # Valida que o match é do mesmo fabricante (CM products only)
                    if kb_row.get("is_cooler_master", "").lower() == "true" or \
                       kb_row.get("marca", "").lower() == "cooler master":
                        matches[pn] = kb_row
                        fuzzy_matches[pn] = f"partial: {kb_pn}"
                        break
    
    exact = sum(1 for p, v in matches.items() if p not in fuzzy_matches)
    fuzzy = len(fuzzy_matches)
    print(f"  Cross-reference CM×KaBuM: {exact} exact + {fuzzy} fuzzy = {len(matches)} total matches")
    return matches

# ─── 8. Excel Builder ────────────────────────────────────────────────────────
def make_header_style(ws, row: int, cols: list, bg: str = COL_HEADER_DARK, fg: str = COL_HEADER_TEXT):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=fg, size=10)
    for i, col_name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_row(ws, row: int, ncols: int, alt: bool = False, bold: bool = False):
    fill = PatternFill("solid", fgColor=COL_ROW_ALT if alt else "FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        if bold:
            cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

def auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len * 1.1))

def freeze_and_filter(ws, freeze_row: int = 2, ncols: int = None):
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    if ncols:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

def status_label(code: str) -> str:
    return {"MP": "Ativo", "EL": "EOL", "BO": "Build Out", "00": "Inativo", "PR": "Pré-Lançamento"}.get(code, code)

def build_excel(pl_rows, kabum_rows, url_map, pl_kb_matches):
    print("\n  Construindo Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Cores por status
    STATUS_FILL = {
        "Ativo":          PatternFill("solid", fgColor="C8E6C9"),
        "EOL":            PatternFill("solid", fgColor="FFCDD2"),
        "Build Out":      PatternFill("solid", fgColor="FFF9C4"),
        "Inativo":        PatternFill("solid", fgColor="ECEFF1"),
        "Pré-Lançamento": PatternFill("solid", fgColor="E1F5FE"),
    }

    # ── Build KaBuM lookup by id
    kb_by_id = {r["id"]: r for r in kabum_rows}
    
    # ── Map pl_row Part No → kabum row
    # (from pl_kb_matches: part_no → kabum_row)

    # ────────────────────────────────────────────────────────────────────────
    # ABA 1: Catálogo Completo (Packing List CM com dados KaBuM)
    # ────────────────────────────────────────────────────────────────────────
    print("    Aba: Catálogo Completo...")
    ws1 = wb.create_sheet("Catálogo Completo")
    cols1 = [
        "BU", "Part No", "Descrição", "Status", "UPC", "EAN",
        "MOQ Pedido", "MOQ Embarque", "PCS/Caixa",
        "IP L(mm)", "IP W(mm)", "IP H(mm)", "IP Peso(kg)",
        "OP L(mm)", "OP W(mm)", "OP H(mm)", "OP Peso(kg)",
        "Caixas/Pallet 20'", "Caixas/Pallet 40'HQ",
        "PCS 20'", "PCS 40'HQ",
        "Link Oficial CM",
        "KaBuM?", "Preço KaBuM (R$)", "Estoque KaBuM", "URL KaBuM",
    ]
    make_header_style(ws1, 1, cols1)
    ws1.row_dimensions[1].height = 30

    for i, pl in enumerate(pl_rows):
        r = i + 2
        pn  = str(pl.get("Part No", "")).strip()
        pn_upper = pn.upper()
        kb  = pl_kb_matches.get(pn_upper)
        kb_id = kb.get("id", "") if kb else ""
        
        status = status_label(str(pl.get("ITEM_STATUS", "")).strip())
        has_kb = "✓ Sim" if kb else "Não"
        
        # Get CM official URL - try from url_map using kabum id, or via slug map
        cm_url = ""
        if kb and kb_id:
            cm_url = url_map.get(kb_id, "")
        
        def n(v):
            try: return float(v) if v else ""
            except: return v or ""

        row_data = [
            pl.get("BU", ""),
            pn,
            pl.get("Item_Desc", ""),
            status,
            str(pl.get("UPC", "")).split(".")[0] if pl.get("UPC") else "",
            str(pl.get("EN_EAN", "")).split(".")[0] if pl.get("EN_EAN") else "",
            n(pl.get("order MOQ", "")),
            n(pl.get("ship MOQ", "")),
            n(pl.get("PCS/ Carton", "")),
            n(pl.get("InnerPacking(L)", "")),
            n(pl.get("InnerPacking(W)", "")),
            n(pl.get("InnerPacking(H)", "")),
            n(pl.get("G.W / PCS", "")),
            n(pl.get("OuterPacking(L)", "")),
            n(pl.get("OuterPacking(W)", "")),
            n(pl.get("OuterPacking(H)", "")),
            n(pl.get("G.W / Carton", "")),
            n(pl.get("Carton / Pallet(20'/40')", "")),
            n(pl.get("Carton / Pallet(40'HQ/45'HQ)", "")),
            n(pl.get("Pallet / PCS(20'/40')", "")),
            n(pl.get("Pallet / PCS(40'HQ/45'HQ)", "")),
            cm_url,
            has_kb,
            float(kb.get("preco_brl", 0)) if kb and kb.get("preco_brl") else "",
            int(kb.get("estoque", 0)) if kb and kb.get("estoque") else "",
            kb.get("url", "") if kb else "",
        ]
        
        for j, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=j, value=val)
            cell.alignment = Alignment(vertical="center")
        
        # Colorir status
        fill = STATUS_FILL.get(status, PatternFill())
        for j in range(1, len(cols1) + 1):
            ws1.cell(row=r, column=j).fill = PatternFill("solid", fgColor=(COL_ROW_ALT if i % 2 else "FFFFFF"))
        ws1.cell(row=r, column=4).fill = fill
        
        # Links clicáveis
        if cm_url:
            ws1.cell(row=r, column=22).hyperlink = cm_url
            ws1.cell(row=r, column=22).font = Font(color="1155CC", underline="single")
        kb_url = kb.get("url", "") if kb else ""
        if kb_url:
            ws1.cell(row=r, column=26).hyperlink = kb_url
            ws1.cell(row=r, column=26).font = Font(color="1155CC", underline="single")

    auto_width(ws1)
    freeze_and_filter(ws1, ncols=len(cols1))
    ws1.column_dimensions["C"].width = 40
    ws1.column_dimensions["V"].width = 55
    ws1.column_dimensions["Z"].width = 55

    # ────────────────────────────────────────────────────────────────────────
    # ABA 2: Todos KaBuM (todos os 403 produtos com dados enriquecidos)
    # ────────────────────────────────────────────────────────────────────────
    print("    Aba: KaBuM Completo...")
    ws2 = wb.create_sheet("KaBuM Completo")
    cols2 = [
        "ID KaBuM", "Marca", "Nome do Produto", "Part Number",
        "Categoria", "Disponível", "Preço (R$)", "Preço PIX (R$)",
        "Preço Lista (R$)", "Desconto (%)", "Estoque",
        "Avaliação", "Nº Avaliações",
        "Peso (g)", "Subcategoria",
        "URL KaBuM", "Link Oficial Fabricante",
        "Data Coleta",
        # CM específico (se for CM)
        "CM BU", "CM Part No Oficial", "CM Descrição Oficial",
        "CM Status", "CM EAN", "CM UPC",
        "CM IP L×W×H (mm)", "CM OP L×W×H (mm)",
        "CM Peso PCS (kg)", "CM PCS/Caixa",
        "CM PCS 20'", "CM PCS 40'HQ",
    ]
    make_header_style(ws2, 1, cols2)
    ws2.row_dimensions[1].height = 30

    # Build reverse map: kabum_id → pl_row
    # We need: for each kabum product that's CM, find PL row
    kb_id_to_pl = {}
    for pl in pl_rows:
        pn_up = str(pl.get("Part No", "")).strip().upper()
        kb = pl_kb_matches.get(pn_up)
        if kb:
            kb_id_to_pl[kb.get("id", "")] = pl

    for i, kb in enumerate(kabum_rows):
        r = i + 2
        kb_id = kb.get("id", "")
        is_cm = kb.get("is_cooler_master", "").strip().lower() == "true"
        oficial_url = url_map.get(kb_id, kb.get("fabricante_url", ""))
        
        pl = kb_id_to_pl.get(kb_id)  # CM PL data if available
        
        def s(k): return kb.get(k, "") or ""
        def f(k):
            try: return float(kb.get(k, 0) or 0) or ""
            except: return ""
        def i_(k):
            try: return int(float(kb.get(k, 0) or 0)) or ""
            except: return ""

        row_data = [
            s("id"), s("marca"), s("nome"), s("partnumber"),
            s("categoria"), "Sim" if s("disponivel").lower()=="true" else "Não",
            f("preco_brl"), f("preco_pix") if "preco_pix" in kb else "",
            f("preco_lista") if "preco_lista" in kb else "",
            f("desconto_pct"), i_("estoque"),
            f("avaliacao"), i_("num_avaliacoes"),
            i_("weight_g"), s("subcategoria"),
            s("url"), oficial_url,
            s("data_coleta"),
            # CM cols
            pl.get("BU", "") if pl else "",
            pl.get("Part No", "") if pl else "",
            pl.get("Item_Desc", "") if pl else "",
            status_label(str(pl.get("ITEM_STATUS", ""))) if pl else "",
            str(pl.get("EN_EAN", "")).split(".")[0] if pl and pl.get("EN_EAN") else "",
            str(pl.get("UPC", "")).split(".")[0] if pl and pl.get("UPC") else "",
            f"{pl.get('InnerPacking(L)','')}×{pl.get('InnerPacking(W)','')}×{pl.get('InnerPacking(H)','')}".replace("0.0×0.0×0.0", "") if pl else "",
            f"{pl.get('OuterPacking(L)','')}×{pl.get('OuterPacking(W)','')}×{pl.get('OuterPacking(H)','')}".replace("0.0×0.0×0.0", "") if pl else "",
            float(pl.get("G.W / PCS", 0) or 0) or "" if pl else "",
            float(pl.get("PCS/ Carton", 0) or 0) or "" if pl else "",
            float(pl.get("Pallet / PCS(20'/40')", 0) or 0) or "" if pl else "",
            float(pl.get("Pallet / PCS(40'HQ/45'HQ)", 0) or 0) or "" if pl else "",
        ]

        for j, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=j, value=val)
            cell.alignment = Alignment(vertical="center")

        # Linha alternada + destaque CM
        alt_fill = PatternFill("solid", fgColor="EBF5FB" if is_cm else (COL_ROW_ALT if i%2 else "FFFFFF"))
        for j in range(1, len(cols2) + 1):
            ws2.cell(row=r, column=j).fill = alt_fill
        
        # Status disponível
        disp_fill = PatternFill("solid", fgColor="C8E6C9") if s("disponivel").lower()=="true" else PatternFill("solid", fgColor="FFCDD2")
        ws2.cell(row=r, column=6).fill = disp_fill

        # Links
        kb_url = s("url")
        if kb_url:
            ws2.cell(row=r, column=16).hyperlink = kb_url
            ws2.cell(row=r, column=16).font = Font(color="1155CC", underline="single")
        if oficial_url and oficial_url.startswith("http"):
            ws2.cell(row=r, column=17).hyperlink = oficial_url
            ws2.cell(row=r, column=17).font = Font(color="1155CC", underline="single")

    auto_width(ws2)
    freeze_and_filter(ws2, ncols=len(cols2))
    ws2.column_dimensions["C"].width = 45
    ws2.column_dimensions["Q"].width = 50

    # ────────────────────────────────────────────────────────────────────────
    # ABA 3: Disponíveis KaBuM (só CM que estão no KaBuM)
    # ────────────────────────────────────────────────────────────────────────
    print("    Aba: CM Disponíveis KaBuM...")
    ws3 = wb.create_sheet("CM Disponíveis KaBuM")
    cols3 = [
        "Part No", "BU", "Descrição Oficial CM", "Status CM",
        "Nome KaBuM", "Preço (R$)", "Preço PIX", "Desconto (%)",
        "Estoque", "Avaliação", "Nº Avaliações",
        "EAN", "UPC",
        "Peso PCS (kg)", "IP Dims (mm)", "OP Dims (mm)",
        "PCS/Caixa", "PCS 20'", "PCS 40'HQ",
        "Link CM Oficial", "URL KaBuM",
    ]
    make_header_style(ws3, 1, cols3)
    ws3.row_dimensions[1].height = 30
    r = 2
    matched_pns = []
    for pl in pl_rows:
        pn_up = str(pl.get("Part No", "")).strip().upper()
        kb = pl_kb_matches.get(pn_up)
        if not kb:
            continue
        matched_pns.append(pn_up)
        kb_id = kb.get("id", "")
        cm_url = url_map.get(kb_id, "")
        status = status_label(str(pl.get("ITEM_STATUS", "")))
        def s(k): return kb.get(k, "") or ""
        def f(k):
            try: return float(kb.get(k, 0) or 0) or ""
            except: return ""
        def i_(k):
            try: return int(float(kb.get(k, 0) or 0)) or ""
            except: return ""

        row_data = [
            pl.get("Part No", ""), pl.get("BU", ""), pl.get("Item_Desc", ""), status,
            s("nome"), f("preco_brl"), f("preco_pix") if "preco_pix" in kb else "",
            f("desconto_pct"), i_("estoque"), f("avaliacao"), i_("num_avaliacoes"),
            str(pl.get("EN_EAN","")).split(".")[0] if pl.get("EN_EAN") else "",
            str(pl.get("UPC","")).split(".")[0] if pl.get("UPC") else "",
            float(pl.get("G.W / PCS", 0) or 0) or "",
            f"{pl.get('InnerPacking(L)','')}×{pl.get('InnerPacking(W)','')}×{pl.get('InnerPacking(H)','')}".replace("0.0×0.0×0.0",""),
            f"{pl.get('OuterPacking(L)','')}×{pl.get('OuterPacking(W)','')}×{pl.get('OuterPacking(H)','')}".replace("0.0×0.0×0.0",""),
            float(pl.get("PCS/ Carton", 0) or 0) or "",
            float(pl.get("Pallet / PCS(20'/40')", 0) or 0) or "",
            float(pl.get("Pallet / PCS(40'HQ/45'HQ)", 0) or 0) or "",
            cm_url, s("url"),
        ]
        for j, val in enumerate(row_data, 1):
            c = ws3.cell(row=r, column=j, value=val)
            c.alignment = Alignment(vertical="center")
        
        fill = STATUS_FILL.get(status, PatternFill("solid", fgColor="FFFFFF"))
        alt_fill = PatternFill("solid", fgColor=COL_ROW_ALT if r%2==0 else "FFFFFF")
        for j in range(1, len(cols3)+1):
            ws3.cell(row=r, column=j).fill = alt_fill
        ws3.cell(row=r, column=4).fill = fill

        if cm_url:
            ws3.cell(row=r, column=20).hyperlink = cm_url
            ws3.cell(row=r, column=20).font = Font(color="1155CC", underline="single")
        kb_url_v = s("url")
        if kb_url_v:
            ws3.cell(row=r, column=21).hyperlink = kb_url_v
            ws3.cell(row=r, column=21).font = Font(color="1155CC", underline="single")
        r += 1

    auto_width(ws3)
    freeze_and_filter(ws3, ncols=len(cols3))
    ws3.column_dimensions["C"].width = 40
    ws3.column_dimensions["E"].width = 45

    # ────────────────────────────────────────────────────────────────────────
    # ABA 4: Não Listados (CM ativos sem KaBuM)
    # ────────────────────────────────────────────────────────────────────────
    print("    Aba: Não Listados CM...")
    ws4 = wb.create_sheet("Não Listados CM")
    cols4 = [
        "BU", "Part No", "Descrição", "Status CM",
        "EAN", "UPC", "MOQ Pedido", "PCS/Caixa",
        "IP L×W×H (mm)", "OP L×W×H (mm)", "Peso PCS (kg)",
        "PCS 20'", "PCS 40'HQ",
        "Link Oficial CM",
    ]
    make_header_style(ws4, 1, cols4)
    ws4.row_dimensions[1].height = 30
    r = 2
    matched_set = set(matched_pns)
    # Add CM products from KaBuM that are NOT in PL (edge case)
    active_statuses = {"MP", "BO", "PR", "00"}
    
    for i, pl in enumerate(pl_rows):
        pn = str(pl.get("Part No", "")).strip()
        pn_up = pn.upper()
        status_code = str(pl.get("ITEM_STATUS", "")).strip()
        if pn_up in matched_set:
            continue
        if status_code not in active_statuses:
            continue
        
        status = status_label(status_code)
        # Doesn't have KaBuM listing → opportunity
        row_data = [
            pl.get("BU", ""), pn, pl.get("Item_Desc", ""), status,
            str(pl.get("EN_EAN","")).split(".")[0] if pl.get("EN_EAN") else "",
            str(pl.get("UPC","")).split(".")[0] if pl.get("UPC") else "",
            float(pl.get("order MOQ",0) or 0) or "",
            float(pl.get("PCS/ Carton",0) or 0) or "",
            f"{pl.get('InnerPacking(L)','')}×{pl.get('InnerPacking(W)','')}×{pl.get('InnerPacking(H)','')}".replace("0.0×0.0×0.0",""),
            f"{pl.get('OuterPacking(L)','')}×{pl.get('OuterPacking(W)','')}×{pl.get('OuterPacking(H)','')}".replace("0.0×0.0×0.0",""),
            float(pl.get("G.W / PCS",0) or 0) or "",
            float(pl.get("Pallet / PCS(20'/40')",0) or 0) or "",
            float(pl.get("Pallet / PCS(40'HQ/45'HQ)",0) or 0) or "",
            "",  # CM URL — preenchido depois se acharmos
        ]
        for j, val in enumerate(row_data, 1):
            c = ws4.cell(row=r, column=j, value=val)
            c.alignment = Alignment(vertical="center")
        
        fill = STATUS_FILL.get(status, PatternFill("solid", fgColor="FFFFFF"))
        alt_fill = PatternFill("solid", fgColor=COL_ROW_ALT if i%2==0 else "FFFFFF")
        for j in range(1, len(cols4)+1):
            ws4.cell(row=r, column=j).fill = alt_fill
        ws4.cell(row=r, column=4).fill = fill
        r += 1

    auto_width(ws4)
    freeze_and_filter(ws4, ncols=len(cols4))
    ws4.column_dimensions["C"].width = 40

    # ────────────────────────────────────────────────────────────────────────
    # ABA 5: EOL
    # ────────────────────────────────────────────────────────────────────────
    print("    Aba: EOL...")
    ws5 = wb.create_sheet("EOL")
    cols5 = ["BU", "Part No", "Descrição", "Status", "EAN", "UPC", "PM_EOL", "KaBuM?", "Preço KaBuM", "URL KaBuM"]
    make_header_style(ws5, 1, cols5)
    ws5.row_dimensions[1].height = 30
    r = 2
    for i, pl in enumerate(pl_rows):
        pn_up = str(pl.get("Part No", "")).strip().upper()
        status_code = str(pl.get("ITEM_STATUS", "")).strip()
        if status_code not in ("EL",):
            continue
        kb = pl_kb_matches.get(pn_up)
        row_data = [
            pl.get("BU",""), pl.get("Part No",""), pl.get("Item_Desc",""),
            status_label(status_code),
            str(pl.get("EN_EAN","")).split(".")[0] if pl.get("EN_EAN") else "",
            str(pl.get("UPC","")).split(".")[0] if pl.get("UPC") else "",
            pl.get("PM_EOL",""),
            "Sim" if kb else "Não",
            float(kb.get("preco_brl",0) or 0) if kb else "",
            kb.get("url","") if kb else "",
        ]
        for j, val in enumerate(row_data, 1):
            c = ws5.cell(row=r, column=j, value=val)
            c.alignment = Alignment(vertical="center")
        fill = PatternFill("solid", fgColor=COL_ROW_ALT if i%2==0 else "FFFFFF")
        for j in range(1, len(cols5)+1):
            ws5.cell(row=r, column=j).fill = fill
        ws5.cell(row=r, column=4).fill = PatternFill("solid", fgColor="FFCDD2")
        if kb and kb.get("url"):
            ws5.cell(row=r, column=10).hyperlink = kb["url"]
            ws5.cell(row=r, column=10).font = Font(color="1155CC", underline="single")
        r += 1
    auto_width(ws5)
    freeze_and_filter(ws5, ncols=len(cols5))
    ws5.column_dimensions["C"].width = 40

    # ────────────────────────────────────────────────────────────────────────
    # ABA 6: Dashboard / Resumo
    # ────────────────────────────────────────────────────────────────────────
    print("    Aba: Dashboard...")
    ws6 = wb.create_sheet("📊 Dashboard", 0)
    ws6.sheet_view.showGridLines = False

    def write_title(ws, row, col, text, size=16, bold=True, color=COL_HEADER_DARK):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center")

    def write_kpi(ws, row, col, label, value, bg=COL_HEADER_DARK, fg=COL_HEADER_TEXT):
        # Label
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = Font(bold=True, size=9, color="888888")
        lc.alignment = Alignment(horizontal="center")
        # Value
        vc = ws.cell(row=row+1, column=col, value=value)
        vc.font = Font(bold=True, size=20, color=bg)
        vc.alignment = Alignment(horizontal="center")
        # Merge
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col)

    # Stats
    total_pl = len(pl_rows)
    active_pl = sum(1 for p in pl_rows if p.get("ITEM_STATUS","") in ("MP","BO","PR","00"))
    eol_pl    = sum(1 for p in pl_rows if p.get("ITEM_STATUS","") == "EL")
    kabum_cm  = len(matched_pns)
    total_kb  = len(kabum_rows)
    kb_avail  = sum(1 for r in kabum_rows if r.get("disponivel","").lower()=="true")
    
    # Brand counts
    brand_counts = Counter(r.get("marca","") for r in kabum_rows)
    
    # BU counts
    bu_counts = Counter(p.get("BU","") for p in pl_rows)
    bu_active = Counter(p.get("BU","") for p in pl_rows if p.get("ITEM_STATUS") in ("MP","BO"))
    bu_kb     = Counter(p.get("BU","") for pn, kb in pl_kb_matches.items()
                        for p in [next((x for x in pl_rows if str(x.get("Part No","")).upper()==pn), None)]
                        if p)

    write_title(ws6, 1, 1, "📊 CM Catálogo Brasil — Dashboard", size=16)
    ws6.row_dimensions[1].height = 32

    write_title(ws6, 3, 1, "KPIs Packing List CM", size=12)
    kpis_pl = [
        ("Total Produtos", total_pl),
        ("Ativos (MP/BO)", active_pl),
        ("EOL", eol_pl),
        ("BR Suffix", sum(1 for p in pl_rows if str(p.get("Part No","")).upper().endswith("BBR"))),
    ]
    for idx, (label, val) in enumerate(kpis_pl):
        write_kpi(ws6, 4, idx+1, label, val)
    
    ws6.row_dimensions[8].height = 10
    write_title(ws6, 9, 1, "KPIs KaBuM Scrape", size=12)
    kpis_kb = [
        ("Total KaBuM", total_kb),
        ("Disponíveis", kb_avail),
        ("CM no KaBuM", kabum_cm),
        ("Cobertura CM %", f"{kabum_cm/max(active_pl,1)*100:.1f}%"),
        ("Marcas", len(brand_counts)),
    ]
    for idx, (label, val) in enumerate(kpis_kb):
        write_kpi(ws6, 10, idx+1, label, val)

    ws6.row_dimensions[14].height = 10
    write_title(ws6, 15, 1, "Distribuição por BU — Packing List", size=12)
    make_header_style(ws6, 16, ["BU","Total","Ativos","No KaBuM","Cobertura %"])
    r = 17
    for bu, total in bu_counts.most_common():
        if not bu:
            continue
        active = bu_active.get(bu, 0)
        in_kb  = bu_kb.get(bu, 0)
        cov    = f"{in_kb/max(active,1)*100:.0f}%" if active else "—"
        row_data = [bu, total, active, in_kb, cov]
        for j, v in enumerate(row_data, 1):
            c = ws6.cell(row=r, column=j, value=v)
            c.alignment = Alignment(horizontal="center" if j>1 else "left")
        style_row(ws6, r, 5, alt=(r%2==0))
        r += 1

    r += 1
    write_title(ws6, r, 1, "Marcas no Scrape KaBuM", size=12)
    r += 1
    make_header_style(ws6, r, ["Marca","Qtd Produtos","Disponíveis","Preço Médio (R$)"])
    r += 1
    for brand, cnt in brand_counts.most_common():
        if not brand: continue
        brand_rows = [x for x in kabum_rows if x.get("marca","") == brand]
        avail = sum(1 for x in brand_rows if x.get("disponivel","").lower()=="true")
        prices = [float(x.get("preco_brl",0) or 0) for x in brand_rows if x.get("preco_brl")]
        avg_price = f"R$ {sum(prices)/len(prices):.0f}" if prices else "—"
        row_data = [brand, cnt, avail, avg_price]
        for j, v in enumerate(row_data, 1):
            c = ws6.cell(row=r, column=j, value=v)
            c.alignment = Alignment(horizontal="center" if j>1 else "left")
        style_row(ws6, r, 4, alt=(r%2==0))
        r += 1

    # Ajustes finais da aba dashboard
    for col in range(1, 7):
        ws6.column_dimensions[get_column_letter(col)].width = 22

    print(f"\n  Salvando Excel: {EXCEL_OUT}")
    wb.save(str(EXCEL_OUT))
    print(f"  ✓ Excel salvo!")
    return {
        "total_pl": total_pl, "active_pl": active_pl, "eol_pl": eol_pl,
        "kabum_cm": kabum_cm, "total_kb": total_kb, "kb_avail": kb_avail,
        "brand_counts": brand_counts, "bu_counts": dict(bu_counts),
        "bu_active": dict(bu_active), "matched_pns": matched_pns,
    }

# ─── 9. HTML Dashboard ────────────────────────────────────────────────────────
def build_html(kabum_rows, url_map, pl_rows, pl_kb_matches, stats):
    print("\n  Construindo Dashboard HTML...")
    
    kb_id_to_pl = {}
    for pl in pl_rows:
        pn_up = str(pl.get("Part No", "")).strip().upper()
        kb = pl_kb_matches.get(pn_up)
        if kb:
            kb_id_to_pl[kb.get("id", "")] = pl

    # Prepare data for JS
    brand_counts = stats["brand_counts"]
    brand_labels = json.dumps([b for b,_ in brand_counts.most_common(15) if b])
    brand_data   = json.dumps([c for b,c in brand_counts.most_common(15) if b])
    
    # Status distribution from PL
    status_map = Counter(status_label(str(p.get("ITEM_STATUS",""))) for p in pl_rows)
    status_labels_js = json.dumps(list(status_map.keys()))
    status_data_js   = json.dumps(list(status_map.values()))
    
    # BU data
    bu_counts = stats["bu_counts"]
    bu_active = stats["bu_active"]
    top_bus = [(bu, cnt) for bu, cnt in sorted(bu_counts.items(), key=lambda x: -x[1]) if bu][:8]
    bu_labels_js = json.dumps([b for b,_ in top_bus])
    bu_total_js  = json.dumps([c for _,c in top_bus])
    bu_active_js = json.dumps([bu_active.get(b,0) for b,_ in top_bus])
    
    # Price distribution
    prices = [float(r.get("preco_brl",0) or 0) for r in kabum_rows if r.get("preco_brl")]
    price_buckets = {"< R$100": 0, "R$100-300": 0, "R$300-600": 0, "R$600-1000": 0, "> R$1000": 0}
    for p in prices:
        if p < 100: price_buckets["< R$100"] += 1
        elif p < 300: price_buckets["R$100-300"] += 1
        elif p < 600: price_buckets["R$300-600"] += 1
        elif p < 1000: price_buckets["R$600-1000"] += 1
        else: price_buckets["> R$1000"] += 1
    price_labels_js = json.dumps(list(price_buckets.keys()))
    price_data_js   = json.dumps(list(price_buckets.values()))
    
    # Table rows
    table_rows = []
    for row in kabum_rows:
        kb_id = row.get("id","")
        is_cm = row.get("is_cooler_master","").strip().lower()=="true"
        pl = kb_id_to_pl.get(kb_id)
        oficial_url = url_map.get(kb_id, row.get("fabricante_url",""))
        
        kb_link = f'<a href="{row.get("url","")}" target="_blank" class="link-kb">KaBuM ↗</a>' if row.get("url") else "—"
        of_link = f'<a href="{oficial_url}" target="_blank" class="link-of">Site ↗</a>' if oficial_url and oficial_url.startswith("http") else "—"
        
        cm_badge = '<span class="badge badge-cm">CM</span>' if is_cm else ""
        
        disp = row.get("disponivel","").lower()=="true"
        disp_badge = '<span class="badge badge-ok">✓</span>' if disp else '<span class="badge badge-no">✗</span>'
        
        price = row.get("preco_brl","")
        try:
            price_fmt = f"R$ {float(price):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        except:
            price_fmt = price or "—"
        
        estoque = row.get("estoque","")
        try: estoque = f"{int(float(estoque)):,}".replace(",",".")
        except: estoque = estoque or "—"
        
        pl_status = f'<span class="status-{status_label(str(pl.get("ITEM_STATUS",""))).lower().replace(" ","-")}">{status_label(str(pl.get("ITEM_STATUS","")))}</span>' if pl else ""
        
        table_rows.append(f"""
        <tr class="{'cm-row' if is_cm else ''}" data-brand="{row.get('marca','')}">
            <td>{row.get('id','')}</td>
            <td class="nome-cell">{cm_badge} {row.get('nome','')[:70]}{'…' if len(row.get('nome',''))>70 else ''}</td>
            <td>{row.get('marca','')}</td>
            <td>{row.get('partnumber','')}</td>
            <td>{row.get('categoria','').split('/')[-1] if row.get('categoria') else ''}</td>
            <td>{disp_badge}</td>
            <td class="price">{price_fmt}</td>
            <td>{estoque}</td>
            <td>{pl_status}</td>
            <td>{kb_link} {of_link}</td>
        </tr>""")

    table_html = "\n".join(table_rows)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CM Catálogo Brasil — Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --bg: #0d1117; --bg2: #161b22; --bg3: #21262d;
    --border: #30363d; --text: #e6edf3; --muted: #8b949e;
    --blue: #58a6ff; --green: #3fb950; --red: #f85149;
    --orange: #d29922; --purple: #bc8cff; --cm-blue: #0066CC;
    --cm-light: #1a3a5c;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         background: var(--bg); color: var(--text); min-height: 100vh; }}
  .header {{
    background: linear-gradient(135deg, var(--cm-blue), var(--cm-light));
    padding: 20px 32px; display: flex; align-items: center; gap: 20px;
    border-bottom: 1px solid var(--border);
    box-shadow: 0 4px 20px rgba(0,102,204,0.3);
  }}
  .header h1 {{ font-size: 22px; font-weight: 700; }}
  .header p {{ font-size: 12px; color: rgba(255,255,255,0.7); margin-top: 2px; }}
  .logo {{ font-size: 32px; }}
  main {{ padding: 24px 32px; max-width: 1600px; margin: 0 auto; }}
  
  /* KPIs */
  .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
               gap: 16px; margin-bottom: 24px; }}
  .kpi-card {{
    background: var(--bg2); border: 1px solid var(--border);
    border-radius: 12px; padding: 20px; text-align: center;
    transition: transform .2s, box-shadow .2s;
  }}
  .kpi-card:hover {{ transform: translateY(-2px); box-shadow: 0 8px 24px rgba(0,0,0,0.4); }}
  .kpi-label {{ font-size: 11px; color: var(--muted); text-transform: uppercase;
                letter-spacing: 1px; margin-bottom: 8px; }}
  .kpi-value {{ font-size: 36px; font-weight: 800; line-height: 1; }}
  .kpi-sub {{ font-size: 11px; color: var(--muted); margin-top: 4px; }}
  .kpi-blue {{ color: var(--blue); }}
  .kpi-green {{ color: var(--green); }}
  .kpi-red {{ color: var(--red); }}
  .kpi-orange {{ color: var(--orange); }}
  .kpi-purple {{ color: var(--purple); }}
  
  /* Charts */
  .charts-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(380px, 1fr));
                  gap: 20px; margin-bottom: 24px; }}
  .chart-card {{
    background: var(--bg2); border: 1px solid var(--border);
    border-radius: 12px; padding: 20px;
  }}
  .chart-card h3 {{ font-size: 14px; font-weight: 600; color: var(--muted);
                    text-transform: uppercase; letter-spacing: .5px; margin-bottom: 16px; }}
  .chart-container {{ position: relative; height: 280px; }}
  
  /* Table */
  .table-section {{
    background: var(--bg2); border: 1px solid var(--border);
    border-radius: 12px; overflow: hidden;
  }}
  .table-header {{
    padding: 16px 20px; display: flex; align-items: center; justify-content: space-between;
    border-bottom: 1px solid var(--border); flex-wrap: wrap; gap: 12px;
  }}
  .table-header h3 {{ font-size: 15px; font-weight: 600; }}
  .controls {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
  input[type=search] {{
    background: var(--bg3); border: 1px solid var(--border); border-radius: 8px;
    color: var(--text); padding: 8px 14px; font-size: 13px; width: 260px;
    outline: none; transition: border .2s;
  }}
  input[type=search]:focus {{ border-color: var(--blue); }}
  select {{
    background: var(--bg3); border: 1px solid var(--border); border-radius: 8px;
    color: var(--text); padding: 8px 12px; font-size: 13px; outline: none;
    cursor: pointer;
  }}
  .table-wrap {{ overflow-x: auto; max-height: 600px; overflow-y: auto; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  thead th {{
    background: var(--bg3); color: var(--muted); font-weight: 600; font-size: 11px;
    text-transform: uppercase; letter-spacing: .5px;
    padding: 10px 12px; text-align: left; position: sticky; top: 0; z-index: 10;
    border-bottom: 1px solid var(--border);
    cursor: pointer; user-select: none;
  }}
  thead th:hover {{ color: var(--text); }}
  tbody tr {{ border-bottom: 1px solid rgba(48,54,61,0.5); transition: background .15s; }}
  tbody tr:hover {{ background: var(--bg3); }}
  tbody tr.cm-row {{ background: rgba(0,102,204,0.06); }}
  tbody tr.cm-row:hover {{ background: rgba(0,102,204,0.12); }}
  td {{ padding: 9px 12px; vertical-align: middle; }}
  .nome-cell {{ max-width: 340px; font-size: 12px; }}
  .price {{ font-weight: 600; color: var(--green); }}
  
  /* Badges */
  .badge {{ display: inline-block; border-radius: 4px; padding: 1px 6px;
            font-size: 10px; font-weight: 700; margin-right: 3px; }}
  .badge-cm {{ background: var(--cm-blue); color: white; }}
  .badge-ok {{ background: rgba(63,185,80,.2); color: var(--green); }}
  .badge-no {{ background: rgba(248,81,73,.2); color: var(--red); }}
  
  /* Status */
  .status-ativo {{ color: var(--green); font-size: 11px; font-weight: 600; }}
  .status-eol {{ color: var(--red); font-size: 11px; }}
  .status-build-out {{ color: var(--orange); font-size: 11px; }}
  .status-inativo {{ color: var(--muted); font-size: 11px; }}
  .status-pré-lançamento {{ color: var(--purple); font-size: 11px; }}
  
  /* Links */
  a {{ color: var(--blue); text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .link-kb {{ color: var(--orange); margin-right: 4px; }}
  .link-of {{ color: var(--blue); }}
  
  /* Footer */
  footer {{ text-align: center; padding: 24px; color: var(--muted); font-size: 12px; }}
  
  /* Divider */
  .section-title {{
    font-size: 13px; font-weight: 600; color: var(--muted);
    text-transform: uppercase; letter-spacing: 1px;
    margin: 24px 0 12px; padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
  }}
</style>
</head>
<body>

<div class="header">
  <div class="logo">🖥️</div>
  <div>
    <h1>Cooler Master — Catálogo Brasil</h1>
    <p>Packing List × KaBuM Scrape · Atualizado em {now}</p>
  </div>
</div>

<main>

  <!-- KPIs Packing List -->
  <div class="section-title">📦 Packing List CM</div>
  <div class="kpi-grid">
    <div class="kpi-card">
      <div class="kpi-label">Total Produtos</div>
      <div class="kpi-value kpi-blue">{stats['total_pl']:,}</div>
      <div class="kpi-sub">no Packing List APAC</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Ativos (MP/BO)</div>
      <div class="kpi-value kpi-green">{stats['active_pl']:,}</div>
      <div class="kpi-sub">{stats['active_pl']/stats['total_pl']*100:.1f}% do catálogo</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">EOL</div>
      <div class="kpi-value kpi-red">{stats['eol_pl']:,}</div>
      <div class="kpi-sub">descontinuados</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">SKUs com BR</div>
      <div class="kpi-value kpi-orange">{sum(1 for p in pl_rows if str(p.get('Part No','')).upper().endswith('BBR')):,}</div>
      <div class="kpi-sub">localização Brasil</div>
    </div>
  </div>

  <!-- KPIs KaBuM -->
  <div class="section-title">🛒 KaBuM Scrape</div>
  <div class="kpi-grid">
    <div class="kpi-card">
      <div class="kpi-label">Total Produtos</div>
      <div class="kpi-value kpi-blue">{stats['total_kb']:,}</div>
      <div class="kpi-sub">scrape {now.split()[0]}</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Disponíveis</div>
      <div class="kpi-value kpi-green">{stats['kb_avail']:,}</div>
      <div class="kpi-sub">{stats['kb_avail']/stats['total_kb']*100:.0f}% em estoque</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">CM no KaBuM</div>
      <div class="kpi-value kpi-purple">{stats['kabum_cm']:,}</div>
      <div class="kpi-sub">produtos matchados</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Cobertura CM</div>
      <div class="kpi-value kpi-orange">{stats['kabum_cm']/max(stats['active_pl'],1)*100:.1f}%</div>
      <div class="kpi-sub">dos ativos no KaBuM</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Marcas</div>
      <div class="kpi-value kpi-blue">{len(stats['brand_counts']):,}</div>
      <div class="kpi-sub">diferentes fabricantes</div>
    </div>
  </div>

  <!-- Charts -->
  <div class="section-title">📊 Análises</div>
  <div class="charts-grid">
    <div class="chart-card">
      <h3>Produtos por Marca (Top 15)</h3>
      <div class="chart-container">
        <canvas id="chartBrands"></canvas>
      </div>
    </div>
    <div class="chart-card">
      <h3>Status Packing List CM</h3>
      <div class="chart-container">
        <canvas id="chartStatus"></canvas>
      </div>
    </div>
    <div class="chart-card">
      <h3>Distribuição por BU (Top 8)</h3>
      <div class="chart-container">
        <canvas id="chartBU"></canvas>
      </div>
    </div>
    <div class="chart-card">
      <h3>Faixa de Preço — KaBuM</h3>
      <div class="chart-container">
        <canvas id="chartPrice"></canvas>
      </div>
    </div>
  </div>

  <!-- Table -->
  <div class="section-title">🔍 Produtos KaBuM (Todos)</div>
  <div class="table-section">
    <div class="table-header">
      <h3>Tabela de Produtos ({len(kabum_rows)} itens)</h3>
      <div class="controls">
        <input type="search" id="searchInput" placeholder="🔍 Buscar produto, marca, SKU..." oninput="filterTable()">
        <select id="brandFilter" onchange="filterTable()">
          <option value="">Todas as marcas</option>
          {"".join(f'<option value="{b}">{b} ({c})</option>' for b,c in brand_counts.most_common() if b)}
        </select>
        <select id="availFilter" onchange="filterTable()">
          <option value="">Disponibilidade</option>
          <option value="sim">✓ Disponível</option>
          <option value="não">✗ Indisponível</option>
        </select>
      </div>
    </div>
    <div class="table-wrap">
      <table id="prodTable">
        <thead>
          <tr>
            <th onclick="sortTable(0)">ID ↕</th>
            <th onclick="sortTable(1)">Produto ↕</th>
            <th onclick="sortTable(2)">Marca ↕</th>
            <th onclick="sortTable(3)">Part Number ↕</th>
            <th onclick="sortTable(4)">Categoria ↕</th>
            <th onclick="sortTable(5)">Disp. ↕</th>
            <th onclick="sortTable(6)">Preço ↕</th>
            <th onclick="sortTable(7)">Estoque ↕</th>
            <th>Status CM</th>
            <th>Links</th>
          </tr>
        </thead>
        <tbody id="tableBody">
          {table_html}
        </tbody>
      </table>
    </div>
    <div style="padding:12px 20px; font-size:12px; color:var(--muted);" id="rowCount">
      Exibindo {len(kabum_rows)} produtos
    </div>
  </div>

</main>

<footer>
  <p>Gerado por Claude Code • Dados: Cooler Master Packing List APAC × KaBuM Scrape • {now}</p>
</footer>

<script>
// ─── Charts ───────────────────────────────────────────────────────────────────
const chartDefaults = {{
  plugins: {{ legend: {{ labels: {{ color: '#8b949e', font: {{ size: 11 }} }} }} }},
  scales: {{
    x: {{ ticks: {{ color: '#8b949e' }}, grid: {{ color: 'rgba(48,54,61,0.6)' }} }},
    y: {{ ticks: {{ color: '#8b949e' }}, grid: {{ color: 'rgba(48,54,61,0.6)' }} }},
  }},
}};

const COLORS15 = ['#58a6ff','#3fb950','#d29922','#bc8cff','#f85149',
                  '#79c0ff','#56d364','#e3b341','#d2a8ff','#ffa198',
                  '#1f6feb','#238636','#9e6a03','#8957e5','#da3633'];

// Brands bar chart
new Chart(document.getElementById('chartBrands'), {{
  type: 'bar',
  data: {{
    labels: {brand_labels},
    datasets: [{{ label: 'Produtos', data: {brand_data},
      backgroundColor: COLORS15, borderRadius: 4 }}]
  }},
  options: {{ ...chartDefaults, plugins: {{ legend: {{ display: false }} }},
    indexAxis: 'y', scales: {{
      x: {{ ticks: {{ color: '#8b949e' }}, grid: {{ color: 'rgba(48,54,61,0.6)' }} }},
      y: {{ ticks: {{ color: '#8b949e' }}, grid: {{ display: false }} }},
    }}
  }}
}});

// Status doughnut
new Chart(document.getElementById('chartStatus'), {{
  type: 'doughnut',
  data: {{
    labels: {status_labels_js},
    datasets: [{{ data: {status_data_js},
      backgroundColor: ['#3fb950','#f85149','#d29922','#8b949e','#bc8cff'],
      borderColor: '#161b22', borderWidth: 2 }}]
  }},
  options: {{ plugins: {{ legend: {{ position: 'right', labels: {{ color: '#8b949e' }} }} }},
    cutout: '60%' }}
}});

// BU grouped bar
new Chart(document.getElementById('chartBU'), {{
  type: 'bar',
  data: {{
    labels: {bu_labels_js},
    datasets: [
      {{ label: 'Total', data: {bu_total_js}, backgroundColor: 'rgba(88,166,255,0.5)',
         borderColor: '#58a6ff', borderWidth: 1, borderRadius: 3 }},
      {{ label: 'Ativos', data: {bu_active_js}, backgroundColor: 'rgba(63,185,80,0.5)',
         borderColor: '#3fb950', borderWidth: 1, borderRadius: 3 }},
    ]
  }},
  options: {{ ...chartDefaults }}
}});

// Price distribution
new Chart(document.getElementById('chartPrice'), {{
  type: 'pie',
  data: {{
    labels: {price_labels_js},
    datasets: [{{ data: {price_data_js},
      backgroundColor: ['#3fb950','#58a6ff','#d29922','#bc8cff','#f85149'],
      borderColor: '#161b22', borderWidth: 2 }}]
  }},
  options: {{ plugins: {{ legend: {{ position: 'right', labels: {{ color: '#8b949e' }} }} }} }}
}});

// ─── Table Filter & Sort ──────────────────────────────────────────────────────
function filterTable() {{
  const q     = document.getElementById('searchInput').value.toLowerCase();
  const brand = document.getElementById('brandFilter').value.toLowerCase();
  const avail = document.getElementById('availFilter').value.toLowerCase();
  const rows  = document.querySelectorAll('#tableBody tr');
  let vis = 0;
  rows.forEach(row => {{
    const text  = row.textContent.toLowerCase();
    const rb    = row.getAttribute('data-brand').toLowerCase();
    const rdis  = row.cells[5].textContent.toLowerCase().includes('✓') ? 'sim' : 'não';
    const show  = (!q || text.includes(q)) &&
                  (!brand || rb === brand) &&
                  (!avail || rdis === avail);
    row.style.display = show ? '' : 'none';
    if (show) vis++;
  }});
  document.getElementById('rowCount').textContent = `Exibindo ${{vis}} produtos`;
}}

let sortDir = {{}};
function sortTable(col) {{
  const tb = document.getElementById('tableBody');
  const rows = Array.from(tb.rows);
  const asc  = !sortDir[col];
  sortDir[col] = asc;
  rows.sort((a, b) => {{
    let av = a.cells[col].textContent.trim();
    let bv = b.cells[col].textContent.trim();
    const an = parseFloat(av.replace(/[R$\s.,]/g,'').replace(',','.'));
    const bn = parseFloat(bv.replace(/[R$\s.,]/g,'').replace(',','.'));
    if (!isNaN(an) && !isNaN(bn)) return asc ? an-bn : bn-an;
    return asc ? av.localeCompare(bv,'pt') : bv.localeCompare(av,'pt');
  }});
  rows.forEach(r => tb.appendChild(r));
}}
</script>
</body>
</html>"""

    HTML_OUT.write_text(html, encoding="utf-8")
    print(f"  ✓ HTML salvo: {HTML_OUT}")

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("CM Catálogo Brasil — Pipeline")
    print("=" * 60)
    
    print("\n[1/6] Carregando dados KaBuM...")
    kabum_rows = load_kabum()
    
    print("\n[2/6] Carregando Packing List CM...")
    pl_rows = load_packing_list()
    
    print("\n[3/6] Baixando sitemaps Cooler Master...")
    cm_slug_map = load_cm_sitemap()
    
    print("\n[4/6] Buscando URLs oficiais dos fabricantes...")
    url_map = build_url_map(kabum_rows, cm_slug_map)
    
    print("\n[5/6] Cruzando Packing List × KaBuM...")
    pl_kb_matches = cross_reference(pl_rows, kabum_rows)
    
    print("\n[6/7] Construindo Excel...")
    stats = build_excel(pl_rows, kabum_rows, url_map, pl_kb_matches)
    stats["pl_rows_ref"] = pl_rows  # pass through for HTML
    
    print("\n[7/7] Construindo Dashboard HTML...")
    build_html(kabum_rows, url_map, pl_rows, pl_kb_matches, stats)
    
    print("\n[+] Copiando para OneDrive...")
    import shutil
    ONEDRIVE_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(EXCEL_OUT), str(ONEDRIVE_DIR / EXCEL_OUT.name))
    shutil.copy2(str(HTML_OUT),  str(ONEDRIVE_DIR / HTML_OUT.name))
    print(f"  ✓ Copiado para {ONEDRIVE_DIR}")
    
    print("\n" + "=" * 60)
    print("✅ CONCLUÍDO!")
    print(f"  Excel:   {EXCEL_OUT}")
    print(f"  HTML:    {HTML_OUT}")
    print(f"  OneDrive: {ONEDRIVE_DIR}")
    print(f"\n  Packing List: {len(pl_rows):,} produtos")
    print(f"  KaBuM: {len(kabum_rows):,} produtos, {len(stats['brand_counts'])} marcas")
    print(f"  CM × KaBuM matches: {len(pl_kb_matches)}")
    print("=" * 60)

if __name__ == "__main__":
    main()
